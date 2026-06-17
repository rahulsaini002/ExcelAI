"""Feature 1.8 — Add a formula column. Verifies 1.8-a..f + edges, including that the
saved .xlsx contains LIVE formulas. Run from backend:  .venv\\Scripts\\python.exe test_1_8.py
"""
from __future__ import annotations

import io

import openpyxl
import pandas as pd

from app.executor import OperationError, execute_plan
from app.main import _serialize

passed = 0
failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
    else:
        failed += 1
        print(f"  FAIL  {name}  {detail}")


def run(op, df):
    return execute_plan(df, [op])


print("FEATURE 1.8 — add a formula column\n")

# 1.8-a multiply two columns -> correct values + LIVE formula in the xlsx
df = pd.DataFrame({"Qty": [2, 3], "Price": [10, 10]})
out, notes, render = run({"action": "add_formula_column", "name": "Total", "formula": "{Qty} * {Price}"}, df)
check("1.8-a values", list(out["Total"]) == [20, 30], str(list(out["Total"])))
check("1.8-a creates named column", "Total" in out.columns)
check("1.8-a emits formula directive", any(d.get("type") == "formula" for d in render))
ws = openpyxl.load_workbook(io.BytesIO(_serialize(out, "o", "xlsx", render)[0])).active
c = ws.cell(row=2, column=3).value  # column C = Total
check("1.8-a live =formula written", isinstance(c, str) and c.startswith("=") and "A2" in c and "B2" in c, repr(c))

# 1.8-b subtraction
out, _, _ = run({"action": "add_formula_column", "name": "Profit", "formula": "{Rev} - {Cost}"}, pd.DataFrame({"Rev": [10, 7], "Cost": [4, 2]}))
check("1.8-b subtraction values", list(out["Profit"]) == [6, 5], str(list(out["Profit"])))

# numbers stored as TEXT still compute numerically
out, _, _ = run({"action": "add_formula_column", "name": "T", "formula": "{a} + {b}"}, pd.DataFrame({"a": ["10", "20"], "b": ["1", "2"]}))
check("text-number columns compute numerically", list(out["T"]) == [11, 22], str(list(out["T"])))

# 1.8-c reference a non-existent column -> friendly error
try:
    run({"action": "add_formula_column", "name": "X", "formula": "{Nope} * 2"}, df)
    check("1.8-c non-existent column caught", False, "no error")
except OperationError as e:
    check("1.8-c non-existent column friendly", "Nope" in str(e))

# 1.8-d division by zero -> no crash; warns; live formula gives #DIV/0! in Excel
out, notes, render = run({"action": "add_formula_column", "name": "D", "formula": "{a} / {b}"}, pd.DataFrame({"a": [10.0, 5.0], "b": [2.0, 0.0]}))
check("1.8-d div-by-zero no crash", "D" in out.columns)
check("1.8-d warns about div-by-zero", "#DIV/0!" in notes[0] or "divide by zero" in notes[0], notes[0])

# 1.8-e new column name already exists -> refuses (asks) unless overwrite
try:
    run({"action": "add_formula_column", "name": "Qty", "formula": "{Qty} * 2"}, df)
    check("1.8-e existing name refused", False, "no error")
except OperationError as e:
    check("1.8-e existing name message", "already exists" in str(e), str(e))
# with overwrite=true it proceeds
out, _, _ = run({"action": "add_formula_column", "name": "Qty", "formula": "{Qty} * 2", "overwrite": True}, df)
check("1.8-e overwrite works", list(out["Qty"]) == [4, 6], str(list(out["Qty"])))

# 1.8-f formula on text columns -> friendly "not numbers" error
try:
    run({"action": "add_formula_column", "name": "Z", "formula": "{Name} * 2"}, pd.DataFrame({"Name": ["a", "b"]}))
    check("1.8-f text columns refused", False, "no error")
except OperationError as e:
    check("1.8-f text columns message", "numbers" in str(e).lower() and "Name" in str(e), str(e))

# --- Excel functions ---
# IF -> conditional column (text results), + live formula written
fdf = pd.DataFrame({"Qty": [1, 3, 5]})
out, notes, render = run({"action": "add_formula_column", "name": "Size", "formula": 'IF({Qty} > 2, "big", "small")'}, fdf)
check("fn IF values", list(out["Size"]) == ["small", "big", "big"], str(list(out["Size"])))
ws = openpyxl.load_workbook(io.BytesIO(_serialize(out, "o", "xlsx", render)[0])).active
c = ws.cell(row=2, column=2).value
check("fn IF live formula", isinstance(c, str) and c.startswith("=IF("), repr(c))

# SUM per row (combine columns)
out, _, _ = run({"action": "add_formula_column", "name": "Tot", "formula": "SUM({a}, {b}, {c})"}, pd.DataFrame({"a": [1, 2], "b": [10, 20], "c": [100, 200]}))
check("fn SUM per row", list(out["Tot"]) == [111, 222], str(list(out["Tot"])))

# AVERAGE per row
out, _, _ = run({"action": "add_formula_column", "name": "Avg", "formula": "AVERAGE({a}, {b})"}, pd.DataFrame({"a": [10, 20], "b": [20, 40]}))
check("fn AVERAGE per row", list(out["Avg"]) == [15.0, 30.0], str(list(out["Avg"])))

# MIN / MAX per row
out, _, _ = run({"action": "add_formula_column", "name": "Lo", "formula": "MIN({a}, {b})"}, pd.DataFrame({"a": [3, 9], "b": [5, 2]}))
check("fn MIN per row", list(out["Lo"]) == [3, 2], str(list(out["Lo"])))
out, _, _ = run({"action": "add_formula_column", "name": "Hi", "formula": "MAX({a}, {b})"}, pd.DataFrame({"a": [3, 9], "b": [5, 2]}))
check("fn MAX per row", list(out["Hi"]) == [5, 9], str(list(out["Hi"])))

# ROUND and ABS
out, _, _ = run({"action": "add_formula_column", "name": "R", "formula": "ROUND({a} / {b}, 2)"}, pd.DataFrame({"a": [10], "b": [3]}))
check("fn ROUND", list(out["R"]) == [3.33], str(list(out["R"])))
out, _, _ = run({"action": "add_formula_column", "name": "AbsD", "formula": "ABS({a} - {b})"}, pd.DataFrame({"a": [2], "b": [9]}))
check("fn ABS", list(out["AbsD"]) == [7], str(list(out["AbsD"])))

# unsupported function -> friendly error, no crash
try:
    run({"action": "add_formula_column", "name": "X", "formula": "VLOOKUP({a})"}, pd.DataFrame({"a": [1]}))
    check("unsupported function errors", False)
except OperationError as e:
    check("unsupported function friendly", "VLOOKUP" in str(e) or "isn't supported" in str(e), str(e))

# safety: a malicious-looking name is not executed (treated as unknown function/name)
try:
    run({"action": "add_formula_column", "name": "X", "formula": "__import__('os')"}, pd.DataFrame({"a": [1]}))
    check("safety: no code execution", False)
except OperationError:
    check("safety: no code execution", True)

# edges
try:
    run({"action": "add_formula_column", "formula": "{Qty} * 2"}, df)
    check("edge: missing name errors", False)
except OperationError:
    check("edge: missing name errors", True)

# parentheses / multiple ops
out, _, _ = run({"action": "add_formula_column", "name": "Net", "formula": "({Rev} - {Cost}) * 2"}, pd.DataFrame({"Rev": [10], "Cost": [3]}))
check("edge: parentheses + precedence", list(out["Net"]) == [14], str(list(out["Net"])))

print(f"\n{passed} passed, {failed} failed.")
raise SystemExit(1 if failed else 0)
