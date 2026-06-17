"""Feature 1.11 — Basic formatting. Verifies 1.11-a..e + acceptance criteria.
Formatting changes DISPLAY only (number_format / bold), never the underlying values.
Checks are done by round-tripping through the .xlsx serializer and reading openpyxl.

Run from backend:  .venv\\Scripts\\python.exe test_1_11.py
"""
from __future__ import annotations

import datetime as dt
import io

import openpyxl
import pandas as pd

from app.executor import OperationError, execute_plan
from app.main import _serialize

passed = 0
failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
    else:
        failed += 1
        print(f"  FAIL  {name}  {detail}")


def run(op, df):
    out, notes, render = execute_plan(df, [op])
    return out, (notes[0] if notes else ""), render


def fmt(**kw):
    base = {"action": "format_cells"}
    base.update(kw)
    return base


def to_ws(df, render):
    return openpyxl.load_workbook(io.BytesIO(_serialize(df, "report", "xlsx", render)[0])).active


def col_idx(df, name):
    return list(df.columns).index(name) + 1


print("FEATURE 1.11 — basic formatting\n")

# --------------------------------------------------------------------------- #
# 1.11-a  Date reformat -> display changes, underlying date unchanged
# --------------------------------------------------------------------------- #
df = pd.DataFrame({"When": pd.to_datetime(["2026-06-09", "2026-12-25"])})
out, note, render = run(fmt(format_columns=["When"], number_format="date"), df)
ws = to_ws(out, render)
c = ws.cell(row=2, column=col_idx(out, "When"))
check("1.11-a date number_format applied", c.number_format == "dd-mm-yyyy", repr(c.number_format))
check("1.11-a underlying date unchanged", c.value == dt.datetime(2026, 6, 9), repr(c.value))

# custom date_format styles
out, _, render = run(fmt(format_columns=["When"], number_format="date", date_format="yyyy-mm-dd"), df)
check("1.11-a custom date_format yyyy-mm-dd", to_ws(out, render).cell(row=2, column=1).number_format == "yyyy-mm-dd")
out, _, render = run(fmt(format_columns=["When"], number_format="date", date_format="dd-mmm-yyyy"), df)
check("1.11-a custom date_format dd-mmm-yyyy", to_ws(out, render).cell(row=2, column=1).number_format == "dd-mmm-yyyy")

# --------------------------------------------------------------------------- #
# 1.11-b  Currency format -> symbol + 2 decimals
# --------------------------------------------------------------------------- #
df = pd.DataFrame({"Amount": [1000, 2500.5]})
out, note, render = run(fmt(format_columns=["Amount"], number_format="currency"), df)
ws = to_ws(out, render)
c = ws.cell(row=2, column=1)
check("1.11-b currency default ₹ + 2dp", c.number_format == '"₹"#,##0.00', repr(c.number_format))
check("1.11-b value unchanged", c.value == 1000, repr(c.value))
# custom symbol + decimals
out, _, render = run(fmt(format_columns=["Amount"], number_format="currency", currency_symbol="$", decimals=0), df)
check("1.11-b custom $ 0dp", to_ws(out, render).cell(row=2, column=1).number_format == '"$"#,##0', repr(None))

# --------------------------------------------------------------------------- #
# 1.11-c  Percentage format -> 0.25 shown as 25%
# --------------------------------------------------------------------------- #
df = pd.DataFrame({"Rate": [0.25, 0.5]})
out, note, render = run(fmt(format_columns=["Rate"], number_format="percent"), df)
ws = to_ws(out, render)
c = ws.cell(row=2, column=1)
check("1.11-c percent format", c.number_format == "0.00%", repr(c.number_format))
check("1.11-c value stays 0.25 (display only)", c.value == 0.25, repr(c.value))
# 0 decimals -> 0%
out, _, render = run(fmt(format_columns=["Rate"], number_format="percent", decimals=0), df)
check("1.11-c percent 0dp", to_ws(out, render).cell(row=2, column=1).number_format == "0%")

# number format
df = pd.DataFrame({"N": [1234.5]})
out, _, render = run(fmt(format_columns=["N"], number_format="number", decimals=2), df)
check("number format thousands", to_ws(out, render).cell(row=2, column=1).number_format == "#,##0.00")

# --------------------------------------------------------------------------- #
# 1.11-d  Bold header row -> header styled, data untouched
# --------------------------------------------------------------------------- #
df = pd.DataFrame({"Name": ["Asha", "Rohan"], "Age": [30, 25]})
out, note, render = run(fmt(bold_header=True), df)
ws = to_ws(out, render)
check("1.11-d header bold", ws.cell(row=1, column=1).font.bold is True and ws.cell(row=1, column=2).font.bold is True)
check("1.11-d data not bold", not ws.cell(row=2, column=1).font.bold)
check("1.11-d data values unchanged", ws.cell(row=2, column=1).value == "Asha" and ws.cell(row=2, column=2).value == 30)

# bold header AND format a column together
out, _, render = run(fmt(format_columns=["Age"], number_format="number", bold_header=True), df)
ws = to_ws(out, render)
check("1.11-d combined bold + format", ws.cell(row=1, column=1).font.bold is True and ws.cell(row=2, column=2).number_format == "#,##0.00")

# --------------------------------------------------------------------------- #
# 1.11-e  Format a column that's the wrong type -> friendly note, no crash
# --------------------------------------------------------------------------- #
df = pd.DataFrame({"Name": ["Asha", "Rohan"]})
out, note, render = run(fmt(format_columns=["Name"], number_format="currency"), df)
check("1.11-e wrong-type friendly note", "Name" in note and "don't look like numbers" in note, note)
check("1.11-e note reassures values unchanged", "weren't changed" in note, note)
ws = to_ws(out, render)  # must not crash
check("1.11-e serializes without crash, value intact", ws.cell(row=2, column=1).value == "Asha")

# date format on free text -> friendly note
out, note, _ = run(fmt(format_columns=["Name"], number_format="date"), df)
check("1.11-e wrong-type date note", "don't look like dates" in note, note)

# numbers-stored-as-text are OK (not flagged) — they're numeric enough
out, note, render = run(fmt(format_columns=["S"], number_format="currency"), pd.DataFrame({"S": ["100", "200"]}))
check("1.11-e numbers-as-text not flagged", "don't look like" not in note, note)

# --------------------------------------------------------------------------- #
# Validation edges
# --------------------------------------------------------------------------- #
# columns given but no format -> asks what format
try:
    run(fmt(format_columns=["Name"]), df)
    check("edge: needs a format", False, "no error")
except OperationError as e:
    check("edge: needs-a-format friendly", "currency" in str(e).lower(), str(e))

# nothing to do (no columns, no bold) -> friendly error
try:
    run(fmt(), df)
    check("edge: nothing to format", False, "no error")
except OperationError as e:
    check("edge: nothing-to-format friendly", "which columns" in str(e).lower() or "bold" in str(e).lower(), str(e))

# non-existent column -> friendly error
try:
    run(fmt(format_columns=["Nope"], number_format="currency"), df)
    check("edge: bad column caught", False, "no error")
except OperationError as e:
    check("edge: bad column friendly", "Nope" in str(e), str(e))

# acceptance: formatting never changes underlying data (df identical to input)
df = pd.DataFrame({"Amount": [1000, 2000]})
out, _, _ = run(fmt(format_columns=["Amount"], number_format="currency"), df)
check("acceptance: data unchanged by formatting", list(out["Amount"]) == [1000, 2000], str(list(out["Amount"])))

print(f"\n{passed} passed, {failed} failed.")
raise SystemExit(1 if failed else 0)
