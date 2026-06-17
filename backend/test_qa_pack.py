"""Recommended QA Dataset Pack — large generated datasets (10k-100k rows) + the
advanced Sort/Filter/Lookup/Formula/Cleaning/Multi-step/Security/Stress cases.
Runs the DETERMINISTIC trusted layer. Brain-judgment cases are marked BRAIN.

Run from backend:  .venv\\Scripts\\python.exe test_qa_pack.py
"""
from __future__ import annotations

import io
import sys
import time

import numpy as np
import openpyxl
import pandas as pd

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from app.executor import execute_multi, execute_plan, OperationError, MultiStepError
from app.reader import load_files
from app.main import _serialize, _serialize_workbook

npass = nfail = nflag = nbrain = 0


def ok(name, cond, got=""):
    global npass, nfail
    if cond:
        npass += 1; print(f"  [PASS] {name}" + (f" — {got}" if got else ""))
    else:
        nfail += 1; print(f"  [FAIL] {name} — {got}")


def flag(name, got):
    global nflag
    nflag += 1; print(f"  [DIFF] {name} — {got}")


def brain(name, note):
    global nbrain
    nbrain += 1; print(f"  [BRAIN] {name} — {note}")


def err(ops, df):
    try:
        execute_plan(df, ops); return None
    except OperationError as e:
        return str(e)


def _safe_load(b):
    try:
        load_files([("x.xlsx", b)]); return None
    except ValueError as e:
        return str(e)
    except Exception as e:
        return f"non-friendly {type(e).__name__}"


rng = np.random.default_rng(7)

# =========================================================================== #
print("\n=== Dataset 1: Sales_Data_Large (10,000 rows) ===")
N = 10_000
regions = ["North", "South", "East", "West"]
sales = pd.DataFrame({
    "Order_ID": range(100001, 100001 + N),
    "Customer_ID": ["C%03d" % (i % 500) for i in range(N)],
    "Customer_Name": rng.choice(["Rahul", "Amit", "Priya", "Sneha", "Rohit", "John"], N),
    "Region": rng.choice(regions, N),
    "Revenue": rng.integers(1000, 50000, N),
    "Cost": rng.integers(500, 30000, N),
    "Date": pd.to_datetime("2025-01-01") + pd.to_timedelta(rng.integers(0, 200, N), "D"),
    "Status": rng.choice(["Completed", "Pending", "Cancelled"], N),
})

# SORT
out, _, _ = execute_plan(sales, [{"action": "sort", "columns": ["Revenue"], "orders": ["desc"]}])
ok("TC-SORT-001 revenue desc", list(out["Revenue"]) == sorted(sales["Revenue"], reverse=True))
out, _, _ = execute_plan(sales, [{"action": "sort", "columns": ["Revenue"], "orders": ["asc"]}])
ok("TC-SORT-002 revenue asc", out["Revenue"].iloc[0] == sales["Revenue"].min())
out, _, _ = execute_plan(sales, [{"action": "sort", "columns": ["Region", "Revenue"], "orders": ["asc", "desc"]}])
nth = out[out["Region"] == "North"]["Revenue"]
ok("TC-SORT-003 region then revenue desc", list(nth) == sorted(nth, reverse=True) and out["Region"].iloc[0] == "East")
out, _, _ = execute_plan(sales, [{"action": "sort", "columns": ["Date"], "orders": ["desc"]}])
ok("TC-SORT-004 date newest first", out["Date"].iloc[0] == sales["Date"].max())
out, _, _ = execute_plan(sales, [{"action": "sort", "columns": ["Customer_Name"], "orders": ["asc"]}])
ok("TC-SORT-005 name A-Z", list(out["Customer_Name"]) == sorted(sales["Customer_Name"], key=str.lower))
work = sales.assign(Profit=sales["Revenue"] - sales["Cost"])
out, _, _ = execute_plan(work, [{"action": "sort", "columns": ["Profit"], "orders": ["desc"]}])
ok("TC-SORT-006 profit highest first", out["Profit"].iloc[0] == work["Profit"].max())
ok("TC-SORT-007 non-existent column errors", (err([{"action": "sort", "columns": ["Ghost"], "orders": ["asc"]}], sales) or "").find("Ghost") >= 0)
out, _, _ = execute_plan(sales, [{"action": "sort", "columns": ["Revenue"], "orders": ["desc"]}, {"action": "limit", "count": 100}])
ok("TC-SORT-008 sort desc + keep top 100", len(out) == 100 and out["Revenue"].iloc[0] == sales["Revenue"].max())

# =========================================================================== #
print("\n=== Filter cases ===")
def f(conds, combine="and"):
    return execute_plan(sales, [{"action": "filter", "conditions": conds, "combine": combine}])[0]
ok("TC-FILTER-001 Revenue > 5000", (f([{"column": "Revenue", "operator": "greater_than", "value": 5000}])["Revenue"] > 5000).all())
ok("TC-FILTER-002 Revenue >= 5000", (f([{"column": "Revenue", "operator": "greater_or_equal", "value": 5000}])["Revenue"] >= 5000).all())
ok("TC-FILTER-003 Revenue < 5000", (f([{"column": "Revenue", "operator": "less_than", "value": 5000}])["Revenue"] < 5000).all())
r = f([{"column": "Revenue", "operator": "between", "value": 10000, "value2": 20000}])["Revenue"]
ok("TC-FILTER-004 between 10k-20k", ((r >= 10000) & (r <= 20000)).all())
ok("TC-FILTER-005 Region = North", (f([{"column": "Region", "operator": "equals", "value": "North"}])["Region"] == "North").all())
ok("TC-FILTER-006 Region != North", (f([{"column": "Region", "operator": "not_equals", "value": "North"}])["Region"] != "North").all())
ok("TC-FILTER-007 Customer contains Rahul", f([{"column": "Customer_Name", "operator": "contains", "value": "Rahul"}])["Customer_Name"].str.contains("Rahul").all())
ok("TC-FILTER-008 Status = Completed", (f([{"column": "Status", "operator": "equals", "value": "Completed"}])["Status"] == "Completed").all())
r = f([{"column": "Status", "operator": "in", "values": ["Completed", "Pending"]}])
ok("TC-FILTER-009 Status in [Completed,Pending]", set(r["Status"]) <= {"Completed", "Pending"} and len(r) > 0)
r = f([{"column": "Date", "operator": "greater_than", "value": "2025-01-01"}])
ok("TC-FILTER-010 Date after Jan 1", (r["Date"] > pd.Timestamp("2025-01-01")).all())
r = f([{"column": "Date", "operator": "between", "value": "2025-01-01", "value2": "2025-03-31"}])
ok("TC-FILTER-011 Date Jan-Mar", ((r["Date"] >= "2025-01-01") & (r["Date"] <= "2025-03-31")).all())
r = f([{"column": "Revenue", "operator": "greater_than", "value": 10000}, {"column": "Region", "operator": "equals", "value": "North"}], "and")
ok("TC-FILTER-012 Rev>10k AND North", ((r["Revenue"] > 10000) & (r["Region"] == "North")).all())
r = f([{"column": "Revenue", "operator": "greater_than", "value": 10000}, {"column": "Region", "operator": "equals", "value": "North"}], "or")
ok("TC-FILTER-013 Rev>10k OR North", ((r["Revenue"] > 10000) | (r["Region"] == "North")).all())
r = f([{"column": "Revenue", "operator": "greater_than", "value": 10000}, {"column": "Region", "operator": "equals", "value": "North"}, {"column": "Status", "operator": "equals", "value": "Completed"}], "and")
ok("TC-FILTER-014 triple AND", len(r) >= 0 and ((r["Revenue"] > 10000) & (r["Region"] == "North") & (r["Status"] == "Completed")).all())
out, _, _ = execute_plan(sales, [{"action": "sort", "columns": ["Revenue"], "orders": ["desc"]}, {"action": "limit", "count": 50}])
ok("TC-FILTER-015 top 50 by revenue", len(out) == 50)

# =========================================================================== #
print("\n=== Lookup (Customers 5k + Orders 50k) ===")
cust = pd.DataFrame({
    "Customer_ID": ["C%04d" % i for i in range(5000)],
    "Customer_Name": ["Name%d" % i for i in range(5000)],
    "Phone": ["9%09d" % i for i in range(5000)],
    "Email": ["u%d@x.com" % i for i in range(5000)],
    "City": rng.choice(["Pune", "Delhi", "Mumbai"], 5000),
})
M = 50_000
orders = pd.DataFrame({
    "Order_ID": range(1, M + 1),
    "Customer_ID": ["C%04d" % i for i in rng.integers(0, 5200, M)],  # some IDs (>=5000) won't match
    "Revenue": rng.integers(100, 9999, M),
})
tables = {"Orders": orders, "Customers": cust}
t = time.time()
out, name, notes, render = execute_multi(tables, "Orders", [{"action": "lookup", "key_column": "Customer_ID", "source_sheet": "Customers", "source_key_column": "Customer_ID", "return_column": "Customer_Name"}])
ok("TC-LKP-001 name lookup 50k", "Customer_Name" in out.columns and len(out) == M, f"{time.time()-t:.2f}s")
spot = out.iloc[0]
ok("TC-LKP-004 exact match correct", spot["Customer_Name"] == ("Not found" if spot["Customer_ID"] not in set(cust["Customer_ID"]) else cust.set_index("Customer_ID").loc[spot["Customer_ID"], "Customer_Name"]))
out2, *_ = execute_multi(tables, "Orders", [{"action": "lookup", "key_column": "Customer_ID", "source_sheet": "Customers", "source_key_column": "Customer_ID", "return_column": "Phone"}])
ok("TC-LKP-002 phone lookup", "Phone" in out2.columns)
out3, *_ = execute_multi(tables, "Orders", [{"action": "lookup", "key_column": "Customer_ID", "source_sheet": "Customers", "source_key_column": "Customer_ID", "return_column": "Email"}])
ok("TC-LKP-003 email lookup", "Email" in out3.columns)
ok("TC-LKP-005 missing id -> 'Not found' (no error)", (out["Customer_Name"] == "Not found").any())
# duplicate source IDs -> first match + note (PRD 1.9); user expected clarification
dup = pd.DataFrame({"Customer_ID": ["C1", "C1"], "Customer_Name": ["First", "Second"]})
o = pd.DataFrame({"Customer_ID": ["C1"]})
_, _, dnotes, _ = execute_multi({"O": o, "S": dup}, "O", [{"action": "lookup", "key_column": "Customer_ID", "source_sheet": "S", "source_key_column": "Customer_ID", "return_column": "Customer_Name"}])
flag("TC-LKP-006 duplicate source IDs", f"current: first-match + note ('{dnotes[0][-60:]}'); user expected a clarification. PRD 1.9 says first-match.")
wb = openpyxl.load_workbook(io.BytesIO(_serialize(out.head(50), "o", "xlsx", render)[0]))
mws = wb["Orders"] if "Orders" in wb.sheetnames else wb.active
fcell = mws.cell(2, list(out.columns).index("Customer_Name") + 1)
ok("TC-LKP-007 XLOOKUP/lookup formula visible", isinstance(fcell.value, str) and fcell.value.startswith("=") and fcell.data_type == "f", repr(fcell.value)[:50])

# =========================================================================== #
print("\n=== Formula cases ===")
fd = pd.DataFrame({"Revenue": [10000, 12000, 15000], "Cost": [5000, 6000, 9000]})
ok("TC-FORM-001 Profit", list(execute_plan(fd, [{"action": "add_formula_column", "name": "Profit", "formula": "{Revenue}-{Cost}"}])[0]["Profit"]) == [5000, 6000, 6000])
ok("TC-FORM-002 Margin %", [round(v) for v in execute_plan(fd, [{"action": "add_formula_column", "name": "M", "formula": "({Revenue}-{Cost})/{Revenue}*100"}])[0]["M"]] == [50, 50, 40])
ok("TC-FORM-003 Tax 18%", [round(v) for v in execute_plan(fd, [{"action": "add_formula_column", "name": "Tax", "formula": "{Revenue}*0.18"}])[0]["Tax"]] == [1800, 2160, 2700])
ok("TC-FORM-004 Net Revenue", list(execute_plan(fd, [{"action": "add_formula_column", "name": "Net", "formula": "{Revenue}-{Cost}"}])[0]["Net"]) == [5000, 6000, 6000])
ok("TC-FORM-005 Discount 10%", [round(v) for v in execute_plan(fd, [{"action": "add_formula_column", "name": "D", "formula": "{Revenue}*0.10"}])[0]["D"]] == [1000, 1200, 1500])
out, _, _ = execute_plan(fd, [{"action": "add_formula_column", "name": "Tier", "formula": 'IF({Revenue} > 10000, "High", "Low")'}])
ok("TC-FORM-006 IF formula", list(out["Tier"]) == ["Low", "High", "High"])
out, _, _ = execute_plan(fd, [{"action": "add_formula_column", "name": "T", "formula": 'IF({Revenue} > 14000, "A", IF({Revenue} > 11000, "B", "C"))'}])
ok("TC-FORM-007 nested IF", list(out["T"]) == ["C", "B", "A"])
brain("TC-FORM-008/009 SUMIFS/COUNTIFS", "cross-row aggregates → use the 'aggregate' op (group_by/count_value), not a per-row formula. Brain routes these.")
ok("TC-FORM-010 formula on non-existent column errors", (err([{"action": "add_formula_column", "name": "X", "formula": "{Ghost}*2"}], fd) or "").find("Ghost") >= 0)

# =========================================================================== #
print("\n=== Dirty data cleaning (10k w/ dupes, blanks, bad types, unicode) ===")
_cust = ["Rahul"] * 500 + ["अमित", "محمد", "张伟", "😀"] * 100 + ["Sneha "] * 100
_cust = (_cust + ["X"] * 10000)[:10000]
_rev = [5000] * 9000 + [""] * 500 + ["ABC", "XYZ", "12A"] * 166 + ["7000", "9000"]
_rev = (_rev + ["0"] * 10000)[:10000]
base = pd.DataFrame({"Customer": _cust, "Revenue": _rev})
out, n, _ = execute_plan(base, [{"action": "remove_duplicates"}])
ok("TC-CLEAN-001 remove duplicates", len(out) < len(base) and "Removed" in n[0], f"{len(base)}->{len(out)}")
out, _, _ = execute_plan(base, [{"action": "fill_missing", "columns": ["Revenue"], "fill_value": "Unknown"}])
ok("TC-CLEAN-002 fill blanks with Unknown", (out["Revenue"] == "Unknown").sum() == 500)
brain("TC-CLEAN-003 standardize date format", "format_cells(date) sets DISPLAY; converting mixed date VALUES to one format isn't a v1 op (note).")
out, n, _ = execute_plan(base, [{"action": "drop_invalid", "columns": ["Revenue"]}])
removed = len(base) - len(out)
ok("TC-CLEAN-004 remove rows with invalid revenue", removed > 0 and pd.to_numeric(out["Revenue"], errors="coerce").isna().sum() == (out["Revenue"] == "").sum(), f"removed {removed} bad rows; note='{n[0][:60]}'")
out, n, _ = execute_plan(pd.DataFrame({"Customer": ["  Rahul ", "Amit  Kumar", "Sneha "]}), [{"action": "trim"}])
ok("TC-CLEAN-005 trim spaces", list(out["Customer"]) == ["Rahul", "Amit Kumar", "Sneha"], str(list(out["Customer"])))

# =========================================================================== #
print("\n=== Multi-step ===")
ms = sales.copy()
out, name, notes, _ = execute_multi({"t": ms}, "t", [
    {"action": "remove_duplicates"},
    {"action": "sort", "columns": ["Revenue"], "orders": ["desc"]},
    {"action": "filter", "conditions": [{"column": "Region", "operator": "equals", "value": "North"}]}])
ok("TC-MS-001 dedupe+sort+filter", (out["Region"] == "North").all() and len(notes) == 3)
out, _, notes, _ = execute_multi({"t": fd}, "t", [
    {"action": "add_formula_column", "name": "Profit", "formula": "{Revenue}-{Cost}"},
    {"action": "filter", "conditions": [{"column": "Profit", "operator": "greater_than", "value": 5000}]},
    {"action": "sort", "columns": ["Profit"], "orders": ["desc"]}])
ok("TC-MS-002 create+filter+sort on new column", (out["Profit"] > 5000).all() and list(out["Profit"]) == sorted(out["Profit"], reverse=True))
out, _, notes, _ = execute_multi(tables, "Orders", [
    {"action": "lookup", "key_column": "Customer_ID", "source_sheet": "Customers", "source_key_column": "Customer_ID", "return_column": "Customer_Name"},
    {"action": "add_formula_column", "name": "Double", "formula": "{Revenue}*2"},
    {"action": "sort", "columns": ["Revenue"], "orders": ["desc"]}])
ok("TC-MS-003 lookup+formula+sort chained", "Customer_Name" in out.columns and "Double" in out.columns)

# =========================================================================== #
print("\n=== Security ===")
ok("Malformed file -> friendly error", isinstance(_safe_load(b"not a spreadsheet"), str))
# prompt injection: the Brain can ONLY emit whitelisted ops; there is NO delete/drop-all op.
from app.llm import Operation
acts = Operation.model_fields["action"].annotation
ok("No destructive op exists (injection can't delete data)", "delete" not in str(acts).lower() and "drop_table" not in str(acts).lower())
brain("Prompt injection ('ignore instructions / delete all')", "structurally safe — Brain outputs only a whitelisted Operation Plan; trusted Python executes. Wording test needs LLM.")

# =========================================================================== #
print("\n=== Stress: 100k rows x 100 cols ===")
SN = 100_000
cols = {"id": range(SN), "Revenue": rng.integers(0, 99999, SN), "Region": rng.choice(regions, SN)}
for c in range(97):
    cols[f"col{c}"] = rng.integers(0, 100, SN)
stress = pd.DataFrame(cols)
t = time.time(); fr = execute_plan(stress, [{"action": "filter", "conditions": [{"column": "Revenue", "operator": "greater_than", "value": 5000}]}])[0]; t_f = time.time() - t
ok("Stress filter 100k x 100", t_f < 5 and len(fr) > 0, f"{t_f:.2f}s, {len(fr)} rows")
t = time.time(); sr = execute_plan(stress, [{"action": "sort", "columns": ["Revenue"], "orders": ["desc"]}])[0]; t_s = time.time() - t
ok("Stress sort 100k x 100", t_s < 5 and sr["Revenue"].iloc[0] == stress["Revenue"].max(), f"{t_s:.2f}s")
t = time.time(); out, _, _ = execute_plan(stress, [{"action": "add_formula_column", "name": "x2", "formula": "{Revenue}*2"}]); t_fm = time.time() - t
ok("Stress formula 100k x 100", out["x2"].iloc[0] == stress["Revenue"].iloc[0] * 2, f"{t_fm:.2f}s")

# =========================================================================== #
print(f"\n================ RESULTS ================")
print(f"PASS: {npass}   FAIL: {nfail}   DIFF: {nflag}   BRAIN (needs live LLM): {nbrain}")
raise SystemExit(1 if nfail else 0)
