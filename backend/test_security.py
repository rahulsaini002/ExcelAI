"""Security & robustness hardening — adversarial "power user" inputs that must NOT
crash, hang, leak, or produce a dangerous file. Complements the per-feature suites.

Run from backend:  .venv\\Scripts\\python.exe test_security.py
"""
from __future__ import annotations

import io
import time

import openpyxl
import pandas as pd
from fastapi.testclient import TestClient

from app import config, main
from app.executor import OperationError, execute_plan
from app.main import _serialize, _serialize_workbook, _too_big, _disarm_injection

passed = failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
    else:
        failed += 1
        print(f"  FAIL  {name}  {detail}")


def run(op, df):
    return execute_plan(df, [op])


print("SECURITY & ROBUSTNESS HARDENING\n")

# --------------------------------------------------------------------------- #
# 1. Safe formula evaluator — no code execution, no DoS
# --------------------------------------------------------------------------- #
# arbitrary code is rejected (AST whitelist)
for bad in ["__import__('os').system('echo hi')", "(1).__class__", "open('x')", "eval('1')"]:
    try:
        run({"action": "add_formula_column", "name": "X", "formula": bad}, pd.DataFrame({"a": [1]}))
        check(f"reject code: {bad[:20]}", False, "no error")
    except OperationError:
        check(f"reject code: {bad[:20]}", True)

# power DoS: huge exponent (constant AND via a column) rejected, fast
for f in ["2 ** 100000", "{a} ** {a}"]:
    t = time.time()
    try:
        run({"action": "add_formula_column", "name": "X", "formula": f}, pd.DataFrame({"a": [99999.0]}))
        check(f"pow DoS blocked: {f}", False, "no error")
    except OperationError:
        check(f"pow DoS blocked: {f}", time.time() - t < 1.0, f"slow: {time.time()-t:.2f}s")
# legitimate small power still works
out, _, _ = run({"action": "add_formula_column", "name": "Sq", "formula": "{a} ** 2"}, pd.DataFrame({"a": [3.0, 4.0]}))
check("small power works", list(out["Sq"]) == [9.0, 16.0], str(list(out["Sq"])))

# overly long formula rejected (defense in depth)
try:
    run({"action": "add_formula_column", "name": "X", "formula": "1" + "+1" * 5000}, pd.DataFrame({"a": [1]}))
    check("long formula rejected", False, "no error")
except OperationError as e:
    check("long formula rejected", "too long" in str(e).lower(), str(e))

# --------------------------------------------------------------------------- #
# 2. Spreadsheet formula INJECTION — uploaded data is never written as a live formula
# --------------------------------------------------------------------------- #
malicious = ['=HYPERLINK("http://evil","click")', "=1+1", "=cmd|'/c calc'!A1", "@SUM(1)", "+1+1", "safe"]
df = pd.DataFrame({"Note": malicious})
ws = openpyxl.load_workbook(io.BytesIO(_serialize(df, "t", "xlsx", [])[0])).active
types = [ws.cell(r, 1).data_type for r in range(2, 2 + len(malicious))]
check("no uploaded cell is a live formula", all(t != "f" for t in types), str(types))
check("injected value preserved as text", ws.cell(2, 1).value == malicious[0], repr(ws.cell(2, 1).value))

# our OWN formulas stay live (disarm doesn't clobber intended formulas)
out, _, render = run({"action": "add_formula_column", "name": "T", "formula": "{q} * {p}"},
                     pd.DataFrame({"q": [2], "p": [5], "Note": ["=evil()"]}))
ws = openpyxl.load_workbook(io.BytesIO(_serialize(out, "t", "xlsx", render)[0])).active
note_idx = list(out.columns).index("Note") + 1
tot_idx = list(out.columns).index("T") + 1
check("intended formula stays live", ws.cell(2, tot_idx).data_type == "f", ws.cell(2, tot_idx).data_type)
check("injected note disarmed alongside", ws.cell(2, note_idx).data_type == "s", ws.cell(2, note_idx).data_type)

# multi-sheet workbook + lookup source are disarmed too
wb = openpyxl.load_workbook(io.BytesIO(_serialize_workbook(
    {"A": pd.DataFrame({"x": ["=evil()"]}), "B": pd.DataFrame({"y": ["=bad()"]})}, "c")[0]))
check("workbook tabs disarmed", all(wb[s].cell(2, 1).data_type == "s" for s in wb.sheetnames), "tab formula leaked")

orders = pd.DataFrame({"CustID": [1]})
people = pd.DataFrame({"ID": [1], "Name": ["=evil()"]})
out, _, render = execute_plan(orders, [{"action": "lookup", "key_column": "CustID", "source_sheet": "People",
                                        "source_key_column": "ID", "return_column": "Name"}],
                              sheets={"Orders": orders, "People": people})
wb = openpyxl.load_workbook(io.BytesIO(_serialize(out, "o", "xlsx", render)[0]))
check("lookup source disarmed", wb["People"].cell(2, 2).data_type == "s", wb["People"].cell(2, 2).data_type)

# disarm helper directly
ws = openpyxl.Workbook().active
ws["A1"] = "=1+1"
_disarm_injection(ws)
check("disarm helper forces text", ws["A1"].data_type == "s")

# --------------------------------------------------------------------------- #
# 3. Upload size limit (DoS guard)
# --------------------------------------------------------------------------- #
class FakeUpload:
    def __init__(self, size):
        self.size = size

limit = config.MAX_UPLOAD_MB
check("under-limit upload allowed", _too_big([FakeUpload(1024)]) is None)
check("over-limit upload rejected", _too_big([FakeUpload((limit + 5) * 1024 * 1024)]) is not None)
check("combined size counted", _too_big([FakeUpload(limit * 1024 * 1024), FakeUpload(2 * 1024 * 1024)]) is not None)
msg = _too_big([FakeUpload((limit + 50) * 1024 * 1024)])
check("size message is friendly", msg and "too large" in msg and str(limit) in msg, msg)

# --------------------------------------------------------------------------- #
# 4. Session memory bounds
# --------------------------------------------------------------------------- #
main._SESSIONS.clear()
for i in range(config.MAX_SESSIONS + 25):
    main._remember_session(f"s{i}", {"tables": {}, "primary": "p", "exts": {}})
check("session count capped", len(main._SESSIONS) <= config.MAX_SESSIONS, str(len(main._SESSIONS)))
check("oldest sessions evicted (newest kept)", f"s{config.MAX_SESSIONS + 24}" in main._SESSIONS)

main._SESSIONS.clear()
main._remember_session("x", {"tables": {}, "primary": "orig", "exts": {}})
for i in range(config.MAX_STATES + 20):
    main._push_state("x", {"tables": {}, "primary": f"step{i}", "exts": {}})
states = main._SESSIONS["x"]["states"]
check("per-session states capped", len(states) <= config.MAX_STATES, str(len(states)))
check("original upload state preserved", states[0]["primary"] == "orig", states[0]["primary"])
check("latest step preserved", states[-1]["primary"] == f"step{config.MAX_STATES + 19}", states[-1]["primary"])
main._SESSIONS.clear()

# --------------------------------------------------------------------------- #
# 5. End-to-end: the /process API never leaks a crash on hostile input
# --------------------------------------------------------------------------- #
client = TestClient(main.app)
_orig = main.llm.parse_instruction
try:
    # a hostile formula routed by a stubbed Brain comes back friendly, not a 500 stack trace
    main.llm.parse_instruction = lambda i, s, h: {"operations": [
        {"action": "add_formula_column", "name": "X", "formula": "2 ** 999999"}]}
    r = client.post("/process", data={"instruction": "boom", "session_id": "z", "rewind": "-1", "history": ""},
                    files=[("files", ("t.csv", b"a\n1\n", "text/csv"))])
    body = r.json()
    check("hostile formula -> friendly error, no crash", r.status_code in (422, 500) and body["status"] == "error", str(body)[:160])
    check("no stack trace leaked", "Traceback" not in r.text and ".py" not in r.text)
finally:
    main.llm.parse_instruction = _orig
    main._SESSIONS.clear()

print(f"\n{passed} passed, {failed} failed.")
raise SystemExit(1 if failed else 0)
