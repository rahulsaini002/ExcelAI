"""Feature 1.13 — Output & download. Verifies the output contract: a valid .xlsx that
opens cleanly, with LIVE formulas and formatting preserved, produced as a NEW file that
never modifies the original. Google-Sheets compatibility is checked via the deterministic
proxy "a well-formed OOXML workbook that openpyxl fully round-trips".

Run from backend:  .venv\\Scripts\\python.exe test_1_13.py
"""
from __future__ import annotations

import copy
import io

import openpyxl
import pandas as pd

from app.executor import execute_multi, execute_plan
from app.main import _serialize, _serialize_workbook, _output_ext

passed = 0
failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
    else:
        failed += 1
        print(f"  FAIL  {name}  {detail}")


def load(out_bytes):
    """Re-open produced bytes as a workbook (raises if corrupt)."""
    return openpyxl.load_workbook(io.BytesIO(out_bytes))


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ZIP_MAGIC = b"PK\x03\x04"  # .xlsx is a zip; this is the OOXML signature


print("FEATURE 1.13 — output & download\n")

# --------------------------------------------------------------------------- #
# 1.13-a  Download after an operation -> valid file, opens in Excel
# --------------------------------------------------------------------------- #
df = pd.DataFrame({"Name": ["Asha", "Rohan"], "Revenue": [300, 100]})
out, notes, render = execute_plan(df, [{"action": "sort", "columns": ["Revenue"], "orders": ["desc"]}])
out_bytes, out_name, media = _serialize(out, "sales.xlsx", "xlsx", render)
check("1.13-a starts with OOXML zip magic", out_bytes[:4] == ZIP_MAGIC)
check("1.13-a media type is xlsx", media == XLSX_MIME, media)
wb = load(out_bytes)  # raises if corrupt
ws = wb.active
check("1.13-a opens & data correct", [ws.cell(1, 1).value, ws.cell(2, 1).value, ws.cell(3, 1).value] == ["Name", "Asha", "Rohan"], "header+sorted rows")

# --------------------------------------------------------------------------- #
# 1.13-b  Formulas preserved -> =B2*C2 is live, recalculates on edit
# --------------------------------------------------------------------------- #
df = pd.DataFrame({"Qty": [2, 3], "Price": [10, 20]})
out, _, render = execute_plan(df, [{"action": "add_formula_column", "name": "Total", "formula": "{Qty} * {Price}"}])
out_bytes, *_ = _serialize(out, "sales.xlsx", "xlsx", render)
ws = load(out_bytes).active
c = ws.cell(row=2, column=3)
check("1.13-b cell holds a LIVE formula", isinstance(c.value, str) and c.value.startswith("="), repr(c.value))
check("1.13-b formula references cells (recalcs)", "A2" in c.value and "B2" in c.value, repr(c.value))
# openpyxl reads it back as a formula (data_type 'f'), confirming it's live not a value
check("1.13-b stored as formula type", c.data_type == "f", c.data_type)

# lookup formula is live too, with the source on its own sheet
orders = pd.DataFrame({"CustID": [1, 2]})
people = pd.DataFrame({"ID": [1, 2], "Name": ["Asha", "Rohan"]})
out, _, render = execute_plan(
    orders,
    [{"action": "lookup", "key_column": "CustID", "source_sheet": "People",
      "source_key_column": "ID", "return_column": "Name"}],
    sheets={"Orders": orders, "People": people},
)
wb = load(_serialize(out, "orders.xlsx", "xlsx", render)[0])
check("1.13-b lookup source sheet present", "People" in wb.sheetnames, str(wb.sheetnames))
lc = wb["Sheet1"].cell(row=2, column=2)
check("1.13-b lookup formula live", isinstance(lc.value, str) and lc.value.startswith("=") and lc.data_type == "f", repr(lc.value))

# --------------------------------------------------------------------------- #
# Formatting preserved -> number_format survives the round-trip
# --------------------------------------------------------------------------- #
df = pd.DataFrame({"Amount": [1000.0, 2500.0]})
out, _, render = execute_plan(df, [{"action": "format_cells", "format_columns": ["Amount"],
                                    "number_format": "currency", "bold_header": True}])
ws = load(_serialize(out, "r.xlsx", "xlsx", render)[0]).active
check("formatting: currency number_format preserved", ws.cell(row=2, column=1).number_format == '"₹"#,##0.00', ws.cell(row=2, column=1).number_format)
check("formatting: bold header preserved", ws.cell(row=1, column=1).font.bold is True)

# highlight (flag_missing) preserved -> blank cell shaded
df = pd.DataFrame({"Region": ["North", None]})
out, _, render = execute_plan(df, [{"action": "flag_missing", "columns": ["Region"]}])
ws = load(_serialize(out, "r.xlsx", "xlsx", render)[0]).active
check("formatting: highlight fill preserved", ws.cell(row=3, column=1).fill.fill_type == "solid", ws.cell(row=3, column=1).fill.fill_type)

# --------------------------------------------------------------------------- #
# 1.13-c  Open in Google Sheets -> opens without corruption (round-trip proxy)
# --------------------------------------------------------------------------- #
# A workbook with formulas + formatting + a second sheet must fully re-open and re-save.
df = pd.DataFrame({"Qty": [2], "Price": [5]})
out, _, render = execute_plan(df, [{"action": "add_formula_column", "name": "Total", "formula": "{Qty}*{Price}"}])
out_bytes, *_ = _serialize(out, "x.xlsx", "xlsx", render)
wb = load(out_bytes)
resaved = io.BytesIO()
wb.save(resaved)  # re-serialising a corrupt file would raise
check("1.13-c re-opens and re-saves cleanly", resaved.getvalue()[:4] == ZIP_MAGIC)

# --------------------------------------------------------------------------- #
# 1.13-d  Multiple operations in a session -> each produces a downloadable result
# --------------------------------------------------------------------------- #
tables = {"sales": pd.DataFrame({"Name": ["A", "B", "A"], "Rev": [10, 20, 10]})}
# step 1: remove duplicates
r1, n1, notes1, render1 = execute_multi(tables, "sales", [{"action": "remove_duplicates"}])
b1 = _serialize(r1, "sales.xlsx", "xlsx", render1)[0]
check("1.13-d step1 dedupe downloadable", load(b1).active.max_row == 3, "header + 2 rows")  # 2 unique rows + header
# step 2: continue from step1's result, add a formula column
r2, n2, notes2, render2 = execute_multi({"sales": r1}, "sales", [{"action": "add_formula_column", "name": "Double", "formula": "{Rev}*2"}])
b2 = _serialize(r2, "sales.xlsx", "xlsx", render2)[0]
ws2 = load(b2).active
check("1.13-d step2 has new formula column", ws2.cell(1, ws2.max_column).value == "Double")
check("1.13-d step2 formula is live", ws2.cell(2, ws2.max_column).data_type == "f")

# --------------------------------------------------------------------------- #
# 1.13-e  Original file integrity -> uploaded original unchanged
# --------------------------------------------------------------------------- #
orig = pd.DataFrame({"Name": ["A", "B", "A"], "Rev": [10, 20, 10]})
tables = {"sales": orig}
before = orig.copy(deep=True)
execute_multi(copy.deepcopy(tables), "sales", [
    {"action": "remove_duplicates"},
    {"action": "sort", "columns": ["Rev"], "orders": ["asc"]},
    {"action": "add_formula_column", "name": "X", "formula": "{Rev}+1"},
])
check("1.13-e input DataFrame not mutated", orig.equals(before), "original changed!")

# the output is a NEW file (never overwrites the original name)
out_bytes, out_name, _ = _serialize(orig, "sales.xlsx", "xlsx", [])
check("1.13-e output is a new filename", out_name == "sales_sumio.xlsx" and out_name != "sales.xlsx", out_name)

# --------------------------------------------------------------------------- #
# CSV can't hold formulas/formatting -> output upgrades to .xlsx (with a note)
# --------------------------------------------------------------------------- #
ext, note = _output_ext("csv", [{"type": "formula"}])
check("csv + formula -> upgrade to xlsx", ext == "xlsx" and note and "preserved" in note, f"{ext} / {note}")
ext, note = _output_ext("csv", [])
check("csv + no directives -> stays csv", ext == "csv" and note is None, f"{ext} / {note}")
ext, note = _output_ext("xlsx", [{"type": "format"}])
check("xlsx stays xlsx", ext == "xlsx" and note is None, f"{ext} / {note}")

# a plain CSV result (no directives) is real CSV with VALUES, not formulas
df = pd.DataFrame({"A": [1, 2]})
out, _, render = execute_plan(df, [{"action": "sort", "columns": ["A"], "orders": ["asc"]}])
csv_bytes, csv_name, csv_media = _serialize(out, "data.csv", "csv", render)
check("csv output is text values", csv_bytes.startswith(b"A") and b"=" not in csv_bytes and csv_name.endswith(".csv"), csv_name)

# --------------------------------------------------------------------------- #
# Multi-sheet workbook (combine_sheets) -> valid file, each tab present
# --------------------------------------------------------------------------- #
sheets = {"Jan": pd.DataFrame({"x": [1]}), "Feb": pd.DataFrame({"y": [2]})}
wb_bytes, wb_name, wb_media = _serialize_workbook(sheets, "combined")
wb = load(wb_bytes)
check("multi-sheet workbook valid", wb_bytes[:4] == ZIP_MAGIC and wb_media == XLSX_MIME)
check("multi-sheet workbook has both tabs", set(wb.sheetnames) == {"Jan", "Feb"}, str(wb.sheetnames))

# --------------------------------------------------------------------------- #
# Download delivery: small results inline as base64; LARGE results are served via
# a streamed /download/{id} URL (no giant base64 in the JSON → no browser OOM).
# --------------------------------------------------------------------------- #
from fastapi.testclient import TestClient  # noqa: E402
from app import main  # noqa: E402

client = TestClient(main.app)
_orig = main.llm.parse_instruction
try:
    main._INLINE_MAX_BYTES = 2000  # tiny threshold so we can test both paths
    main.llm.parse_instruction = lambda i, s, h: {"operations": [{"action": "sort", "columns": ["A"], "orders": ["asc"]}]}

    # small result -> inline base64 present AND a download_id
    small = client.post("/process", data={"instruction": "sort", "session_id": "d1", "rewind": "-1", "history": ""},
                        files=[("files", ("t.csv", b"A\n3\n1\n2\n", "text/csv"))]).json()
    check("small result inlines base64", bool(small.get("file_base64")) and bool(small.get("download_id")), str(small)[:120])

    # large result -> NO inline base64, but a download_id, and the file streams from /download
    big_csv = b"A\n" + b"\n".join(str(i).encode() for i in range(5000))
    big = client.post("/process", data={"instruction": "sort", "session_id": "d2", "rewind": "-1", "history": ""},
                      files=[("files", ("big.csv", big_csv, "text/csv"))]).json()
    check("large result has NO inline base64", big.get("file_base64") in (None, "") and bool(big.get("download_id")), f"b64={bool(big.get('file_base64'))}")
    dl = client.get(f"/download/{big['download_id']}")
    check("download endpoint streams the file", dl.status_code == 200 and dl.content.startswith(b"A") and len(dl.content) > 2000, f"{dl.status_code}, {len(dl.content)} bytes")
    check("download has attachment filename", "attachment" in dl.headers.get("content-disposition", ""), dl.headers.get("content-disposition"))
    check("missing download id -> 404", client.get("/download/nope").status_code == 404)
    # HEAD lets the frontend check a file exists before downloading (history downloads)
    check("HEAD existing -> 200", client.head(f"/download/{big['download_id']}").status_code == 200)
    check("HEAD missing -> 404", client.head("/download/nope").status_code == 404)

    # DURABILITY (option B): the result file survives a backend restart. Simulate one
    # by wiping the in-memory index and reloading it from disk, then download again.
    rid = big["download_id"]
    main._RESULTS.clear()                # as if the process restarted (memory gone)
    main._load_results_index()           # re-attach to files on disk
    dl2 = client.get(f"/download/{rid}")
    check("result survives a restart (reloaded from disk)", dl2.status_code == 200 and dl2.content.startswith(b"A"), f"{dl2.status_code}")
finally:
    main.llm.parse_instruction = _orig
    main._INLINE_MAX_BYTES = 6 * 1024 * 1024
    main._SESSIONS.clear()
    # tidy the test's result files off disk
    for _rid in list(main._RESULTS):
        main._delete_result(_rid)
    main._save_results_index()

print(f"\n{passed} passed, {failed} failed.")
raise SystemExit(1 if failed else 0)
