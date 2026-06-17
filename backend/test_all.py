"""Comprehensive test of the trusted logic (reader + executor) against the PRD's
test cases and extra edge cases. No API key needed.

Run from the backend folder:  .venv\\Scripts\\python.exe test_all.py
"""
from __future__ import annotations

import io

import openpyxl
import pandas as pd

from app.executor import MultiStepError, OperationError, execute_multi, execute_plan
from app.main import _serialize, _serialize_workbook
from app.reader import load_files

passed = 0
failed = 0
fails: list[str] = []


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
    else:
        failed += 1
        fails.append(f"{name}  {detail}")
        print(f"  FAIL  {name}  {detail}")


def run(ops, df, sheets=None):
    return execute_plan(df, ops, sheets)


def _err(ops, df, sheets=None):
    """True if the plan raises a clean OperationError (not some other crash)."""
    try:
        execute_plan(df, ops, sheets)
        return False
    except OperationError:
        return True
    except Exception:
        return False


def _err_multi(tables, primary, ops):
    try:
        execute_multi(tables, primary, ops)
        return False
    except OperationError:
        return True
    except Exception:
        return False


def xlsx_bytes(rows: list[list]) -> bytes:
    """Build a raw .xlsx from a list of cell rows (for layout-specific tests)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def csv_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_csv(buf, index=False)
    return buf.getvalue()


# ====================================================================== READER
print("READER")

# 1.1-b CSV
ld = load_files([("data.csv", csv_bytes(pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})))])
check("1.1-b csv loads", list(ld.tables["data"].columns) == ["A", "B"])

# phantom blank leading columns dropped (testing.xlsx pattern)
b = xlsx_bytes([[None, None, "product", "Quantity"], [None, None, 1001, 1], [None, None, 1111, 2]])
ld = load_files([("testing.xlsx", b)])
check("phantom empty Unnamed cols dropped", list(ld.tables["testing"].columns) == ["product", "Quantity"],
      str(list(ld.tables["testing"].columns)))

# title row above headers -> promote (Participant.xlsx pattern)
b = xlsx_bytes([["Competition", None, None], ["Name", "Subject", "Grade"], ["Quiz", "Eng", "IX"], ["Debate", "Hindi", "X"]])
ld = load_files([("part.xlsx", b)])
check("1.1 title-row header recovered", list(ld.tables["part"].columns) == ["Name", "Subject", "Grade"],
      str(list(ld.tables["part"].columns)))

# 1.1-e duplicate headers -> disambiguated by pandas
ld = load_files([("dup.csv", b"Amount,Amount\n1,2\n3,4\n")])
cols = list(ld.tables["dup"].columns)
check("1.1-e duplicate headers disambiguated", cols == ["Amount", "Amount.1"], str(cols))

# 1.1-k non-English (Hindi) headers preserved
ld = load_files([("hin.csv", "नाम,मूल्य\nA,10\n".encode("utf-8"))])
check("1.1-k non-English headers preserved", list(ld.tables["hin"].columns) == ["नाम", "मूल्य"],
      str(list(ld.tables["hin"].columns)))

# 1.1-h wrong file type rejected
try:
    load_files([("pic.pdf", b"%PDF-1.4 junk")])
    check("1.1-h wrong type rejected", False, "no error")
except ValueError:
    check("1.1-h wrong type rejected", True)

# 1.1-g empty-ish file (headers only, no rows) doesn't crash
ld = load_files([("empty.csv", b"A,B\n")])
check("1.1-g header-only file loads", len(ld.tables["empty"]) == 0)

# multi-sheet workbook -> one table per sheet
wb = openpyxl.Workbook()
wb.active.title = "First"
wb.active.append(["x", "y"]); wb.active.append([1, 2])
s2 = wb.create_sheet("Second"); s2.append(["p"]); s2.append([9])
buf = io.BytesIO(); wb.save(buf)
ld = load_files([("multi.xlsx", buf.getvalue())])
check("multi-sheet -> table per sheet", set(ld.tables) == {"multi - First", "multi - Second"}, str(list(ld.tables)))

# ======================================================================= SORT
print("SORT")
check("1.4-a numbers sort numerically",
      list(run([{"action": "sort", "columns": ["n"]}], pd.DataFrame({"n": [10, 2, 1]}))[0]["n"]) == [1, 2, 10])
check("1.4-a numbers-as-text sort numerically",
      list(run([{"action": "sort", "columns": ["n"]}], pd.DataFrame({"n": ["10", "2", "1"]}))[0]["n"]) == ["1", "2", "10"],
      "text-numbers should sort 1,2,10 not 1,10,2")
check("1.4-b descending",
      list(run([{"action": "sort", "columns": ["n"], "orders": ["desc"]}], pd.DataFrame({"n": [1, 3, 2]}))[0]["n"]) == [3, 2, 1])
check("1.4-d text sorts case-insensitively",
      list(run([{"action": "sort", "columns": ["t"]}], pd.DataFrame({"t": ["banana", "apple", "Cherry"]}))[0]["t"]) == ["apple", "banana", "Cherry"],
      "expected apple,banana,Cherry")
check("1.4-c dates sort chronologically",
      list(run([{"action": "sort", "columns": ["d"]}], pd.DataFrame({"d": pd.to_datetime(["2022-05-01", "2020-01-01", "2021-03-03"])}))[0]["d"].dt.year) == [2020, 2021, 2022])
df_blank = pd.DataFrame({"n": [3, None, 1]})
check("1.4-f blanks sorted to end",
      pd.isna(run([{"action": "sort", "columns": ["n"]}], df_blank)[0]["n"].iloc[-1]))
check("sort missing column errors", _err([{"action": "sort", "columns": ["nope"]}], pd.DataFrame({"n": [1]})))

# ===================================================================== FILTER
print("FILTER")
sample = pd.DataFrame({"Region": ["North", "South", ""], "Price": [100, 200, 50], "When": pd.to_datetime(["2021-01-01", "2022-06-01", "2020-03-03"])})
check("1.5-a text equals", len(run([{"action": "filter", "conditions": [{"column": "Region", "operator": "equals", "value": "North"}]}], sample)[0]) == 1)
check("1.5-b number greater_than", len(run([{"action": "filter", "conditions": [{"column": "Price", "operator": "greater_than", "value": "100"}]}], sample)[0]) == 1)
check("1.5-c date after",
      len(run([{"action": "filter", "conditions": [{"column": "When", "operator": "greater_than", "value": "2021-06-01"}]}], sample)[0]) == 1,
      "date comparison should work")
check("1.5-d contains", len(run([{"action": "filter", "conditions": [{"column": "Region", "operator": "contains", "value": "out"}]}], sample)[0]) == 1)
check("1.5-e AND of two", len(run([{"action": "filter", "combine": "and", "conditions": [{"column": "Region", "operator": "equals", "value": "North"}, {"column": "Price", "operator": "greater_than", "value": "50"}]}], sample)[0]) == 1)
check("1.5-f matches nothing -> 0 rows", len(run([{"action": "filter", "conditions": [{"column": "Region", "operator": "equals", "value": "Mars"}]}], sample)[0]) == 0)
check("1.5-h is_blank", len(run([{"action": "filter", "conditions": [{"column": "Region", "operator": "is_blank"}]}], sample)[0]) == 1)
check("between numeric", len(run([{"action": "filter", "conditions": [{"column": "Price", "operator": "between", "value": "60", "value2": "150"}]}], sample)[0]) == 1)
check("1.5-i wrong-type compare errors", _err([{"action": "filter", "conditions": [{"column": "Region", "operator": "greater_than", "value": "5"}]}], sample))

# =========================================================== REMOVE DUPLICATES
print("REMOVE DUPLICATES")
check("1.6-a exact dup removed", len(run([{"action": "remove_duplicates"}], pd.DataFrame({"a": [1, 1, 2], "b": ["x", "x", "y"]}))[0]) == 2)
check("1.6-b by key column", len(run([{"action": "remove_duplicates", "columns": ["Email"]}], pd.DataFrame({"Email": ["a@x.com", "a@x.com"], "n": [1, 2]}))[0]) == 1)
check("1.6-c none present", len(run([{"action": "remove_duplicates", "columns": ["Email"]}], pd.DataFrame({"Email": ["a", "b"]}))[0]) == 2)
check("1.6-d case/space treated same",
      len(run([{"action": "remove_duplicates", "columns": ["Email"]}], pd.DataFrame({"Email": ["a@x.com ", "A@X.com"]}))[0]) == 1,
      "trimmed + case-insensitive should collapse to 1")

# ============================================================ MISSING VALUES
print("MISSING VALUES")
check("1.7-a fill blanks", (run([{"action": "fill_missing", "columns": ["R"], "fill_value": "U"}], pd.DataFrame({"R": ["x", None, ""]}))[0]["R"] == "U").sum() == 2)
check("1.7-b drop rows with blanks", len(run([{"action": "drop_missing", "columns": ["R"]}], pd.DataFrame({"R": ["x", None, ""]}))[0]) == 1)
check("1.7-a fill numeric keeps number", run([{"action": "fill_missing", "columns": ["n"], "fill_value": "0"}], pd.DataFrame({"n": [1.0, None]}))[0]["n"].iloc[1] == 0)

# ============================================================ FORMULA COLUMN
print("FORMULA COLUMN")
check("1.8-a multiply", list(run([{"action": "add_formula_column", "name": "T", "formula": "{q} * {p}"}], pd.DataFrame({"q": [2, 3], "p": [10, 10]}))[0]["T"]) == [20, 30])
check("1.8-b subtract", list(run([{"action": "add_formula_column", "name": "P", "formula": "{rev} - {cost}"}], pd.DataFrame({"rev": [10], "cost": [4]}))[0]["P"]) == [6])
check("1.8-c missing col errors", _err([{"action": "add_formula_column", "name": "X", "formula": "{nope} * 2"}], pd.DataFrame({"q": [1]})))
try:
    run([{"action": "add_formula_column", "name": "D", "formula": "{a} / {b}"}], pd.DataFrame({"a": [1.0], "b": [0.0]}))
    check("1.8-d div-by-zero no crash", True)
except OperationError:
    check("1.8-d div-by-zero no crash", True)  # a clean error is also acceptable
except Exception as e:
    check("1.8-d div-by-zero no crash", False, str(e))

# ===================================================================== LOOKUP
print("LOOKUP")
main = pd.DataFrame({"id": [1, 2, 3]})
ref = pd.DataFrame({"id": [1, 2], "name": ["A", "B"]})
out = run([{"action": "lookup", "key_column": "id", "source_sheet": "ref", "source_key_column": "id", "return_column": "name"}], main, {"m": main, "ref": ref})[0]
check("1.9-a cross-table lookup", list(out["name"]) == ["A", "B", "Not found"])
check("1.9-e source missing errors", _err([{"action": "lookup", "key_column": "id", "source_sheet": "ghost", "source_key_column": "id", "return_column": "name"}], main, {"m": main}))

# ================================================================== AGGREGATE
print("AGGREGATE")
agg = pd.DataFrame({"Region": ["N", "S", "N"], "Price": [100, 200, 50], "Note": ["a", "b", "c"]})
# scalar aggregates report the value in the note and keep the data
check("1.10-a sum", ": 350" in run([{"action": "aggregate", "agg_func": "sum", "agg_column": "Price"}], agg)[1][0])
check("1.10-c count", ": 3" in run([{"action": "aggregate", "agg_func": "count", "agg_column": "Region"}], agg)[1][0])
g = run([{"action": "aggregate", "agg_func": "sum", "agg_column": "Price", "group_by": ["Region"]}], agg)[0]
check("1.10-d grouped sum", set(g["sum_of_Price"]) == {150, 200})
check("1.10-f sum text errors", _err([{"action": "aggregate", "agg_func": "sum", "agg_column": "Note"}], agg))
check("1.10-b average ignores blanks", ": 15" in run([{"action": "aggregate", "agg_func": "average", "agg_column": "v"}], pd.DataFrame({"v": [10, None, 20]}))[1][0])

# =============================================================== FIND/REPLACE
print("FIND/REPLACE")
check("1.12-a simple replace", list(run([{"action": "find_replace", "column": "c", "find": "mumbai", "replace": "Mumbai"}], pd.DataFrame({"c": ["mumbai", "Delhi"]}))[0]["c"]) == ["Mumbai", "Delhi"])
check("1.12-b case-insensitive", (run([{"action": "find_replace", "column": "c", "find": "x", "replace": "Y"}], pd.DataFrame({"c": ["X", "x"]}))[0]["c"] == "Y").sum() == 2)
check("1.12-c whole-cell only", list(run([{"action": "find_replace", "column": "c", "find": "cat", "replace": "dog", "whole_cell": True}], pd.DataFrame({"c": ["cat", "category"]}))[0]["c"]) == ["dog", "category"])
check("1.12-d not found message", "No cells" in run([{"action": "find_replace", "column": "c", "find": "zzz", "replace": "q"}], pd.DataFrame({"c": ["a"]}))[1][0])

# ================================================================= COLUMN OPS
print("COLUMN OPS")
cdf = pd.DataFrame({"A": [1], "B": [2], "C": [3]})
check("drop_columns", list(run([{"action": "drop_columns", "columns": ["B"]}], cdf)[0].columns) == ["A", "C"])
check("select_columns", list(run([{"action": "select_columns", "columns": ["C", "A"]}], cdf)[0].columns) == ["C", "A"])
check("rename_columns", "Q" in run([{"action": "rename_columns", "rename_from": ["A"], "rename_to": ["Q"]}], cdf)[0].columns)

# ====================================================================== MERGE
print("MERGE")
t1 = pd.DataFrame({"Customer_ID": [1], "Amt": [10]})
t2 = pd.DataFrame({"client_id": [2], "Amt": [20]})
m, name, notes, _ = execute_multi({"t1": t1, "t2": t2}, "t1", [{"action": "merge", "merge_tables": ["t1", "t2"], "column_groups": [{"name": "Customer_ID", "aliases": ["client_id"]}]}])
check("merge synonym groups unify", "Customer_ID" in m.columns and "client_id" not in m.columns and len(m) == 2, str(list(m.columns)))
p = pd.DataFrame({"Customer ID": [1]})
q = pd.DataFrame({"customer_id": [2]})
m2, *_ = execute_multi({"p": p, "q": q}, "p", [{"action": "merge", "merge_tables": ["p", "q"]}])
check("merge auto-unify case/space", len(m2.columns) == 1 and len(m2) == 2, str(list(m2.columns)))

# merge of UNRELATED tables (no shared columns) -> side by side, no staircase
d1 = pd.DataFrame({"Name": ["A", "B"], "Roll": [1, 2]})
d2 = pd.DataFrame({"product": [10, 20, 30]})
md, *_ = execute_multi({"d1": d1, "d2": d2}, "d1", [{"action": "merge", "merge_tables": ["d1", "d2"]}])
check("merge disjoint -> side by side", list(md.columns) == ["Name", "Roll", "product"] and len(md) == 3, str(list(md.columns)) + f" rows={len(md)}")
check("merge disjoint puts product on row 1", md["product"].iloc[0] == 10, str(list(md["product"])))

# combine_sheets -> ONE workbook with each table on its OWN tab (not stacked)
csres, csname, csnotes, _ = execute_multi(
    {"Jan": pd.DataFrame({"x": [1, 2]}), "Feb": pd.DataFrame({"y": [3]})}, "Jan",
    [{"action": "combine_sheets", "sheet_tables": ["Jan", "Feb"], "new_table": "book"}])
check("combine_sheets returns multi-sheet dict", isinstance(csres, dict) and set(csres) == {"Jan", "Feb"}, str(type(csres)))
cwb = openpyxl.load_workbook(io.BytesIO(_serialize_workbook(csres, csname)[0]))
check("combine_sheets writes separate tabs", set(cwb.sheetnames) == {"Jan", "Feb"}, str(cwb.sheetnames))
check("combine_sheets <2 tables errors", _err_multi({"A": pd.DataFrame({"x": [1]})}, "A", [{"action": "combine_sheets", "sheet_tables": ["A"]}]))

# ================================================================ MULTI-STEP
print("MULTI-STEP")
ms = pd.DataFrame({"R": ["N", "S", "N"], "P": [3, 1, 2]})
out = execute_multi({"t": ms}, "t", [{"action": "filter", "conditions": [{"column": "R", "operator": "equals", "value": "N"}]}, {"action": "sort", "columns": ["P"], "orders": ["desc"]}])[0]
check("MS-a two steps in order", list(out["P"]) == [3, 2])
# MS-b: a bad SECOND step keeps the first step's result and reports the failure.
try:
    execute_multi({"t": ms}, "t", [{"action": "sort", "columns": ["P"], "orders": ["desc"]}, {"action": "drop_columns", "columns": ["ghost"]}])
    check("MS-b later-step failure is partial", False, "no error")
except MultiStepError as e:
    check("MS-b later-step failure is partial",
          e.failed_step == 2 and list(e.partial_result["P"]) == [3, 2, 1] and "ghost" in e.reason,
          f"step={e.failed_step} reason={e.reason[:40]}")
# a bad FIRST step has no partial result -> plain friendly error
check("MS-b first-step failure is a plain error", _err_multi({"t": ms}, "t", [{"action": "drop_columns", "columns": ["ghost"]}, {"action": "sort", "columns": ["P"]}]))

# ========================================= OUTPUT: live formulas / highlight / .xls
print("OUTPUT (formulas / highlight / corrupted)")

# add_formula_column emits a live-formula directive...
fdf, _, render = run([{"action": "add_formula_column", "name": "T", "formula": "{q} * {p}"}],
                     pd.DataFrame({"q": [2, 3], "p": [10, 10]}))
check("1.8 formula emits a formula directive", any(d.get("type") == "formula" for d in render), str(render))
check("1.8 formula still computes values", list(fdf["T"]) == [20, 30])

# ...and the saved .xlsx contains a LIVE =formula (not just the value)
xbytes = _serialize(fdf, "out", "xlsx", render)[0]
wb = openpyxl.load_workbook(io.BytesIO(xbytes))
ws = wb.active
c2 = ws.cell(row=2, column=3).value  # column C = "T"
check("1.8/1.13 live Excel formula written", isinstance(c2, str) and c2.startswith("=") and "A2" in c2 and "B2" in c2, repr(c2))

# flag_missing highlights blanks without changing data
hdf, hnotes, hrender = run([{"action": "flag_missing", "columns": ["R"]}], pd.DataFrame({"R": ["x", None, ""]}))
check("1.7-c flag emits highlight directive", any(d.get("type") == "highlight" for d in hrender))
check("1.7-c flag leaves data unchanged", list(hdf["R"].fillna("")) == ["x", "", ""])
hb = _serialize(hdf, "out", "xlsx", hrender)[0]
hws = openpyxl.load_workbook(io.BytesIO(hb)).active
check("1.7-c blank cell highlighted", hws.cell(row=3, column=1).fill.fill_type == "solid")
check("1.7-c non-blank cell not highlighted", hws.cell(row=2, column=1).fill.fill_type in (None, "none"))

# 1.9 lookup -> live =XLOOKUP + source written as its own sheet
mainf = pd.DataFrame({"id": [1, 2, 3]})
reff = pd.DataFrame({"id": [1, 2], "name": ["A", "B"]})
ldf, lnotes, lrender = run([{"action": "lookup", "key_column": "id", "source_sheet": "ref", "source_key_column": "id", "return_column": "name"}], mainf, {"m": mainf, "ref": reff})
check("1.9 lookup emits lookup directive", any(d.get("type") == "lookup" for d in lrender))
lwb = openpyxl.load_workbook(io.BytesIO(_serialize(ldf, "out", "xlsx", lrender)[0]))
check("1.9 source written as its own sheet", len(lwb.sheetnames) >= 2, str(lwb.sheetnames))
# default style = INDEX/MATCH (works in every Excel version + Google Sheets)
import app.config as cfg
nameval = lwb["Sheet1"].cell(row=2, column=2).value
check("1.9 default lookup uses INDEX/MATCH", isinstance(nameval, str) and "INDEX(" in nameval and "MATCH(" in nameval, repr(nameval))
# xlookup style when configured
cfg.LOOKUP_STYLE = "xlookup"
valx = openpyxl.load_workbook(io.BytesIO(_serialize(ldf, "out", "xlsx", lrender)[0]))["Sheet1"].cell(row=2, column=2).value
check("1.9 XLOOKUP style when configured", isinstance(valx, str) and valx.startswith("=XLOOKUP("), repr(valx))
cfg.LOOKUP_STYLE = "index_match"

# 1.1-i corrupted file -> friendly ValueError, no crash
try:
    load_files([("broken.xlsx", b"this is definitely not a real xlsx file")])
    check("1.1-i corrupted file friendly error", False, "no error")
except ValueError:
    check("1.1-i corrupted file friendly error", True)
except Exception as e:
    check("1.1-i corrupted file friendly error", False, f"wrong type {type(e).__name__}")


print(f"\n{passed} passed, {failed} failed.")
if fails:
    print("FAILURES:")
    for f in fails:
        print("  -", f)
raise SystemExit(1 if failed else 0)
