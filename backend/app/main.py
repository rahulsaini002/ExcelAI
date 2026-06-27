"""FastAPI app: the engine that ties everything together.

Flow for POST /process:
  1. read the uploaded file (pandas)
  2. summarize its structure
  3. call Gemini -> operation plan (or a clarifying question)
  4. execute the plan (pandas)
  5. return the processed file (base64) + a plain-language "here's what I did"
"""
from __future__ import annotations

import base64
import io
import json
import re
import time
import traceback
import uuid
from collections import OrderedDict
from pathlib import Path

import pandas as pd
from fastapi import FastAPI, File, Form, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Matches {ColumnName} placeholders inside a formula template.
_PLACEHOLDER = re.compile(r"\{([^{}]+)\}")

from . import config, fallback, llm
from .executor import MultiStepError, OperationError, execute_multi
from .reader import load_files, summarize_structure, summarize_tables

app = FastAPI(title="Sumio API", version="0.1.0")

# Open for local dev (any localhost/127.0.0.1/LAN origin) so the browser is never
# blocked by CORS. In production set SUMIO_CORS_ALLOW_ALL=0 to restrict to the origins
# listed in SUMIO_CORS_ORIGINS (your deployed frontend).
if config.CORS_ALLOW_ALL:
    app.add_middleware(
        CORSMiddleware,
        allow_origin_regex=".*",
        allow_methods=["*"],
        allow_headers=["*"],
    )
else:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=config.CORS_ORIGINS,
        allow_methods=["*"],
        allow_headers=["*"],
    )


# In-memory per-session working data, keyed by a session id from the frontend.
# Each entry: {"tables": {name: DataFrame}, "primary": str, "exts": {name: ext}}.
# This lets follow-up instructions build on the previous result (chaining) instead
# of re-reading the original uploads every time. Cleared when the server restarts.
_SESSIONS: dict[str, dict] = {}

# Generated result files, served on demand via GET /download/{id}. We hand the
# frontend a small download id instead of inlining the whole file as base64 in the
# JSON (base64-in-JSON makes the browser hold several copies of a big file and run out
# of memory). Files are written to DISK + indexed in memory, so downloads — and the
# frontend's "continue on the result" — SURVIVE A BACKEND RESTART (even for big files).
# Bounded by total size and a TTL.
_RESULTS_DIR = Path(config.RESULTS_DIR)
_RESULTS_INDEX = _RESULTS_DIR / "index.json"
_RESULTS_MAX_BYTES = config.MAX_RESULTS_MB * 1024 * 1024
_RESULTS_TTL = config.RESULTS_TTL_HOURS * 3600
_RESULTS: "OrderedDict[str, dict]" = OrderedDict()  # id -> {filename, media_type, size, created}
# Results small enough to ALSO inline as base64 (instant download for normal files).
_INLINE_MAX_BYTES = 6 * 1024 * 1024


def _save_results_index() -> None:
    try:
        _RESULTS_INDEX.write_text(json.dumps(_RESULTS))
    except Exception:
        pass


def _delete_result(rid: str) -> None:
    _RESULTS.pop(rid, None)
    try:
        (_RESULTS_DIR / rid).unlink(missing_ok=True)
    except Exception:
        pass


def _prune_results() -> None:
    """Drop expired results, then the oldest until under the size cap."""
    now = time.time()
    for rid in list(_RESULTS):
        if now - _RESULTS[rid].get("created", 0) > _RESULTS_TTL:
            _delete_result(rid)
    total = sum(m["size"] for m in _RESULTS.values())
    while total > _RESULTS_MAX_BYTES and len(_RESULTS) > 1:
        rid, meta = next(iter(_RESULTS.items()))
        total -= meta["size"]
        _delete_result(rid)


def _load_results_index() -> None:
    """Re-attach to result files written before a restart (so downloads still work)."""
    _RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    try:
        for rid, meta in json.loads(_RESULTS_INDEX.read_text()).items():
            if (_RESULTS_DIR / rid).exists():
                _RESULTS[rid] = meta
    except Exception:
        pass
    _prune_results()
    _save_results_index()


def _store_result(out_bytes: bytes, filename: str, media_type: str) -> str:
    """Write a result to disk for download and return its id (evicts old ones)."""
    rid = uuid.uuid4().hex
    try:
        (_RESULTS_DIR / rid).write_bytes(out_bytes)
    except Exception:
        traceback.print_exc()
    _RESULTS[rid] = {
        "filename": filename, "media_type": media_type,
        "size": len(out_bytes), "created": time.time(),
    }
    _prune_results()
    _save_results_index()
    return rid


_load_results_index()


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "model": config.MODEL}


def _abbrev(x: float) -> str:
    """Compact human number: 4,820,000 -> 4.82M, 18204 -> 18.2K."""
    ax = abs(x)
    if ax >= 1_000_000:
        return f"{x / 1_000_000:.2f}M"
    if ax >= 1_000:
        return f"{x / 1_000:.1f}K"
    return f"{int(x)}" if x == int(x) else f"{x:.2f}"


def _format_number(x: float, fmt: str | None) -> str:
    if fmt == "percent":
        return f"{x:.1f}%"
    if fmt == "currency":
        return "₹" + _abbrev(x)
    if abs(x) >= 10_000:
        return _abbrev(x)
    return f"{int(x):,}" if x == int(x) else f"{x:,.2f}"


def _compute_kpi(df: pd.DataFrame, metric: dict) -> str | None:
    """Compute a single KPI value from the data, formatted for display."""
    agg = metric.get("agg")
    col = metric.get("column")
    fmt = metric.get("format")
    try:
        if agg == "count":
            return _format_number(float(len(df)), fmt or "number")
        if not col or col not in df.columns:
            return None
        if agg == "count_distinct":
            return _format_number(float(df[col].nunique()), fmt or "number")
        series = pd.to_numeric(df[col], errors="coerce").dropna()
        if series.empty:
            return None
        if agg == "sum":
            val = float(series.sum())
        elif agg in ("mean", "average"):
            val = float(series.mean())
        elif agg == "min":
            val = float(series.min())
        elif agg == "max":
            val = float(series.max())
        else:
            return None
        return _format_number(val, fmt)
    except Exception:
        return None


def _compute_series(df: pd.DataFrame, metric: dict, top: int = 8) -> list[float]:
    """Compute a numeric series (an aggregate per group) for a chart."""
    agg = metric.get("agg")
    col = metric.get("column")
    gb = metric.get("group_by")
    try:
        if not gb or gb not in df.columns:
            return []
        if agg == "count" or not col or col not in df.columns:
            grouped = df.groupby(gb).size()
        else:
            vals = pd.to_numeric(df[col], errors="coerce")
            tmp = pd.DataFrame({"_g": df[gb].values, "_v": vals.values}).dropna(subset=["_v"])
            g = tmp.groupby("_g")["_v"]
            grouped = {
                "sum": g.sum, "mean": g.mean, "average": g.mean, "min": g.min, "max": g.max,
            }.get(agg, g.sum)()
        grouped = grouped.sort_values(ascending=False).head(top)
        return [round(float(v), 2) for v in grouped.tolist()]
    except Exception:
        return []


def _compute_table(df: pd.DataFrame, metric: dict, top: int = 10) -> dict | None:
    """Compute a small aggregated table (group_by + aggregate) for a report block."""
    agg = metric.get("agg")
    col = metric.get("column")
    gb = metric.get("group_by")
    fmt = metric.get("format")
    try:
        if not gb or gb not in df.columns:
            return None
        if agg == "count" or not col or col not in df.columns:
            grouped = df.groupby(gb).size()
            value_label = "Count"
        else:
            vals = pd.to_numeric(df[col], errors="coerce")
            tmp = pd.DataFrame({"_g": df[gb].values, "_v": vals.values}).dropna(subset=["_v"])
            g = tmp.groupby("_g")["_v"]
            grouped = {
                "sum": g.sum, "mean": g.mean, "average": g.mean, "min": g.min, "max": g.max,
            }.get(agg, g.sum)()
            value_label = f"{str(agg).title()} {col}"
        grouped = grouped.sort_values(ascending=False).head(top)
        rows = [[str(idx), _format_number(float(v), fmt)] for idx, v in grouped.items()]
        return {"columns": [str(gb), value_label], "rows": rows}
    except Exception:
        return None


@app.post("/dashboard")
async def dashboard(
    prompt: str = Form(...),
    columns: str = Form(""),
    files: list[UploadFile] = File(default=[]),
) -> JSONResponse:
    """Design a dashboard (a set of widgets) from a plain-language prompt. The Brain
    chooses the widgets + a metric (agg + column) for each; if a DATA FILE is provided,
    trusted code then computes the REAL numbers from it. On model failure the frontend
    falls back to a local template, so we return a clear error here."""
    prompt = (prompt or "").strip()
    if not prompt:
        return _error("Please describe the dashboard you'd like.", status=400)

    # With a data file we compute real numbers; otherwise use the columns hint only.
    df: pd.DataFrame | None = None
    if files:
        too_big = _too_big(files)
        if too_big:
            return _error(too_big, status=413)
        uploads = [(f.filename or "upload", await f.read()) for f in files]
        try:
            data = load_files(uploads)
        except ValueError as exc:
            return _error(str(exc), status=400)
        except Exception:
            return _error(_INTERNAL_ERROR, status=500)
        df = data.tables[data.primary]
        structure = summarize_tables(data.tables, data.primary)
    else:
        # `columns` is an optional JSON array like [{"name": "...", "type": "..."}].
        try:
            structure = json.loads(columns) if columns.strip() else {}
        except Exception:
            structure = {}

    try:
        spec = llm.generate_dashboard(prompt, structure)
    except Exception as exc:
        unavailable = isinstance(exc, llm.ModelUnavailableError)
        key_missing = isinstance(exc, RuntimeError) and not unavailable
        if not (unavailable or key_missing):
            traceback.print_exc()
        if unavailable:
            return _error(str(exc), status=503)
        if key_missing:
            return _error(str(exc), status=500)
        return _error(
            "I couldn't reach the AI service to build the dashboard right now — "
            "please try again in a moment.",
            status=502,
        )

    # Fill in REAL numbers from the data wherever the model gave a metric.
    if df is not None:
        for w in spec.get("widgets", []):
            metric = w.get("metric")
            if not metric:
                continue
            if w.get("type") == "kpi":
                val = _compute_kpi(df, metric)
                if val is not None:
                    w["value"] = val
                    w["delta"] = None  # real value — no fabricated change
            elif w.get("type") == "chart" and metric.get("group_by"):
                series = _compute_series(df, metric)
                if series:
                    w["data"] = series
        spec["computed"] = True

    return JSONResponse({"status": "ok", **spec})


def _safe_filename(name: str) -> str:
    """A safe download filename stem from a report title."""
    return re.sub(r"[^\w\-]+", "_", (name or "").strip()).strip("_")[:40] or "report"


def _build_report_xlsx(title: str, source: str, blocks: list[dict]) -> bytes:
    """Render a report definition (title + ordered blocks) into a formatted .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    meta = [f"Source: {source}"] if source else []
    meta.append(time.strftime("%d %b %Y"))
    ws["A2"] = " · ".join(meta)
    ws["A2"].font = Font(italic=True, color="888888")

    row = 4
    for b in blocks:
        bt = b.get("type")
        btitle = (b.get("title") or "").strip()
        if bt == "kpi":
            ws.cell(row=row, column=1, value=btitle or "Metric").font = Font(bold=True)
            ws.cell(row=row, column=2, value=b.get("value") or "")
            if b.get("delta"):
                ws.cell(row=row, column=3, value=b.get("delta"))
            row += 1
        elif bt == "narrative":
            ws.cell(row=row, column=1, value=btitle or "Narrative").font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value=b.get("text") or "")
            row += 2
        elif bt == "chart":
            ws.cell(row=row, column=1, value=btitle or "Chart").font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"[{b.get('chartType') or 'chart'} chart]")
            row += 2
        elif bt == "table":
            ws.cell(row=row, column=1, value=btitle or "Table").font = Font(bold=True)
            row += 1
            cols = b.get("columns") or []
            for ci, col in enumerate(cols, start=1):
                ws.cell(row=row, column=ci, value=col).font = Font(bold=True)
            if cols:
                row += 1
            for r in b.get("rows") or []:
                for ci, val in enumerate(r, start=1):
                    ws.cell(row=row, column=ci, value=val)
                row += 1
            row += 1
        else:
            row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    _disarm_injection(ws)  # the report text is user-influenced — neutralize "=" cells
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/report/export")
async def report_export(report: str = Form(...)) -> JSONResponse:
    """Build a formatted .xlsx from a report definition (JSON: title, source, blocks)
    and return it via the same download mechanism as processed files."""
    try:
        data = json.loads(report)
    except Exception:
        return _error("That report couldn't be read — please try again.", status=400)

    title = (data.get("title") or "Report").strip() or "Report"
    source = (data.get("source") or "").strip()
    blocks = data.get("blocks") or []
    try:
        out_bytes = _build_report_xlsx(title, source, blocks)
    except Exception:
        traceback.print_exc()
        return _error(_INTERNAL_ERROR, status=500)

    filename = f"{_safe_filename(title)}.xlsx"
    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    download_id = _store_result(out_bytes, filename, media)
    inline = (
        base64.b64encode(out_bytes).decode("ascii")
        if len(out_bytes) <= _INLINE_MAX_BYTES else None
    )
    return JSONResponse({
        "status": "ok",
        "filename": filename,
        "media_type": media,
        "download_id": download_id,
        "file_base64": inline,
    })


@app.post("/report/compute")
async def report_compute(
    blocks: str = Form(...),
    files: list[UploadFile] = File(default=[]),
) -> JSONResponse:
    """Bind a report's blocks to a data file: the Brain assigns a metric per block, then
    trusted code computes real KPI values, chart series, and table rows from the file."""
    try:
        block_list = json.loads(blocks)
    except Exception:
        return _error("That report couldn't be read — please try again.", status=400)
    if not isinstance(block_list, list) or not block_list:
        return _error("This report has no blocks to compute.", status=400)
    if not files:
        return _error("Please choose a data file to compute from.", status=400)

    too_big = _too_big(files)
    if too_big:
        return _error(too_big, status=413)
    uploads = [(f.filename or "upload", await f.read()) for f in files]
    try:
        data = load_files(uploads)
    except ValueError as exc:
        return _error(str(exc), status=400)
    except Exception:
        return _error(_INTERNAL_ERROR, status=500)
    df = data.tables[data.primary]
    structure = summarize_tables(data.tables, data.primary)

    try:
        plan = llm.assign_report_metrics(block_list, structure)
    except Exception as exc:
        unavailable = isinstance(exc, llm.ModelUnavailableError)
        key_missing = isinstance(exc, RuntimeError) and not unavailable
        if not (unavailable or key_missing):
            traceback.print_exc()
        if unavailable:
            return _error(str(exc), status=503)
        if key_missing:
            return _error(str(exc), status=500)
        return _error(
            "I couldn't reach the AI service to compute the report — please try again.",
            status=502,
        )

    metrics = {
        it["index"]: it.get("metric")
        for it in plan.get("items", [])
        if isinstance(it.get("index"), int) and it.get("metric")
    }
    for i, b in enumerate(block_list):
        metric = metrics.get(i)
        if not metric:
            continue
        t = b.get("type")
        if t == "kpi":
            v = _compute_kpi(df, metric)
            if v is not None:
                b["value"] = v
                b["delta"] = None
        elif t == "chart" and metric.get("group_by"):
            s = _compute_series(df, metric)
            if s:
                b["data"] = s
        elif t == "table" and metric.get("group_by"):
            tbl = _compute_table(df, metric)
            if tbl:
                b["columns"] = tbl["columns"]
                b["rows"] = tbl["rows"]

    return JSONResponse({"status": "ok", "blocks": block_list})


@app.api_route("/download/{result_id}", methods=["GET", "HEAD"])
def download(result_id: str):
    """Stream a generated result file from disk (the browser saves it straight to disk,
    so a large file never lives in the page's memory; survives a server restart).
    HEAD is supported so the frontend can check a file still exists before downloading."""
    meta = _RESULTS.get(result_id)
    path = _RESULTS_DIR / result_id
    if not meta or not path.exists():
        return _error("That download has expired — please re-run the step.", status=404)
    return FileResponse(path, media_type=meta["media_type"], filename=meta["filename"])


@app.post("/undo")
async def undo(session_id: str = Form(...)) -> JSONResponse:
    """Undo the last successful step: pop the most recent state so the next instruction
    builds on the step before it. Returns how many steps remain."""
    entry = _SESSIONS.get(session_id)
    states = entry["states"] if entry else []
    if len(states) <= 1:  # states[0] is the original upload — nothing to undo
        return _error("There's nothing to undo yet.", status=400)
    states.pop()
    cur = states[-1]
    biggest = max((len(t) for t in cur["tables"].values()), default=0)
    return JSONResponse({
        "status": "ok",
        "steps_remaining": len(states) - 1,  # not counting the original upload
        "row_count": biggest,
        "primary": cur["primary"],
    })


@app.post("/inspect")
async def inspect(
    files: list[UploadFile] = File(...),
    session_id: str = Form(""),
) -> JSONResponse:
    """Read uploaded file(s) and return their structure (sheets, columns + types,
    row count, sample rows) so the UI can show a preview BEFORE any operation.

    If a `session_id` is given, the loaded data is ALSO remembered for that session
    so the two-phase flow (/parse then /execute) can reuse it without re-uploading."""
    if not files:
        return _error("Please upload a spreadsheet.", status=400)
    too_big = _too_big(files)
    if too_big:
        return _error(too_big, status=413)
    uploads = [(f.filename or "upload", await f.read()) for f in files]
    try:
        data = load_files(uploads)
    except ValueError as exc:
        return _error(str(exc), status=400)
    except Exception:
        return _error(_INTERNAL_ERROR, status=500)

    # Remember the upload for the session so /parse + /execute can use it.
    if session_id:
        _remember_session(
            session_id,
            {"tables": dict(data.tables), "primary": data.primary, "exts": dict(data.exts)},
        )

    tables = []
    for name, df in data.tables.items():
        s = summarize_structure(df, sample_rows=5)
        rc = s["row_count"]
        note = None
        if rc == 0:
            note = "This sheet has no data rows."
        elif rc > 50_000:
            note = f"Large file ({rc:,} rows) — preview shows the first 5 rows."
        tables.append(
            {
                "name": name,
                "row_count": rc,
                "columns": s["columns"],
                "sample_rows": s["sample_rows"],
                "note": note,
            }
        )
    return JSONResponse({"status": "ok", "tables": tables})


def _describe_plan(operations: list[dict]) -> str:
    """A deterministic one-line plain-language restatement of a plan, used when the
    model didn't provide its own 'translation' (e.g. the offline fallback parser)."""
    parts: list[str] = []
    for op in operations:
        a = op.get("action")
        cols = ", ".join(op.get("columns") or [])
        if a == "sort":
            order = (op.get("orders") or ["asc"])[0]
            parts.append(
                f"sort by {cols or 'the chosen column'} "
                f"({'high to low' if order == 'desc' else 'low to high'})"
            )
        elif a == "filter":
            parts.append("keep only the rows matching your condition")
        elif a == "limit":
            parts.append(f"keep the {'last' if op.get('from_end') else 'top'} {op.get('count') or 'N'} rows")
        elif a == "remove_duplicates":
            parts.append("remove duplicate rows" + (f" on {cols}" if cols else ""))
        elif a == "add_formula_column":
            parts.append(f"add a '{op.get('name') or 'new'}' column")
        elif a == "aggregate":
            parts.append(f"{op.get('agg_func') or 'aggregate'} {op.get('agg_column') or ''}".strip())
        elif a == "lookup":
            parts.append(f"look up {op.get('return_column') or 'a value'} from another table")
        elif a == "merge":
            parts.append("merge the tables")
        elif a in ("fill_missing", "drop_missing", "drop_invalid", "flag_missing"):
            parts.append(a.replace("_", " ") + (f" in {cols}" if cols else ""))
        else:
            parts.append((a or "operation").replace("_", " "))
    return ", then ".join(parts) if parts else "run the operation"


@app.post("/parse")
async def parse(
    instruction: str = Form(...),
    session_id: str = Form(""),
    history: str = Form(""),
) -> JSONResponse:
    """Phase 1 of the two-phase flow: the Brain ONLY. Translate the instruction into an
    operation plan WITHOUT executing it, so the UI can preview the interpretation +
    confidence before running. Ambiguous -> clarify; unsupported -> message; the file is
    never touched. /execute then runs the returned plan."""
    instruction = (instruction or "").strip()
    if not instruction:
        return _error("Please describe what you'd like done to the data.", status=400)

    entry = _SESSIONS.get(session_id) if session_id else None
    if not entry or not entry.get("states"):
        return _error("Please upload a spreadsheet to start.", status=400)
    base = entry["states"][-1]
    tables, primary = base["tables"], base["primary"]
    structure = summarize_tables(tables, primary)

    # Translate via the Brain (falling back to the deterministic parser if it's down).
    try:
        plan = llm.parse_instruction(instruction, structure, history)
    except Exception as exc:
        unavailable = isinstance(exc, llm.ModelUnavailableError)
        key_missing = isinstance(exc, RuntimeError) and not unavailable
        if not (unavailable or key_missing):
            traceback.print_exc()
        plan = fallback.parse(instruction, structure)
        if plan is None:
            if unavailable:
                return _error(str(exc), status=503)
            if key_missing:
                return _error(str(exc), status=500)
            return _error(
                "I couldn't reach the AI service to understand your request right now — "
                "this isn't a problem with your instruction. Please try again in a moment.",
                status=502,
            )

    clarification = plan.get("clarification")
    reply = plan.get("reply")
    operations = plan.get("operations") or []
    if not operations:
        if reply:
            return JSONResponse({"status": "message", "message": reply})
        if clarification:
            return JSONResponse({"status": "clarify", "clarification": clarification})
        return JSONResponse(
            {
                "status": "message",
                "message": (
                    "I didn't understand that — try describing the task, e.g. "
                    '"sort by Revenue descending" or "remove duplicate rows".'
                ),
            }
        )

    translation = (plan.get("translation") or "").strip() or _describe_plan(operations)
    confidence = plan.get("confidence")
    if not isinstance(confidence, int) or not (0 <= confidence <= 100):
        confidence = 80
    return JSONResponse(
        {
            "status": "plan",
            "translation": translation,
            "confidence": confidence,
            # The full plan the UI hands back to /execute (no second Brain call).
            "plan": {"operations": operations, "title": (plan.get("title") or "").strip() or None},
        }
    )


def _run_operations(
    session_id: str, base: dict, operations: list[dict], ai_title, started_at: float
) -> JSONResponse:
    """Run an operation plan on a base state, push the new state, serialize, and build
    the OK response. Shared shape with /process (deltas, formulas, preview, partial
    warnings, streamed download). Returns a friendly 422 on an expected step failure or
    500 on an unexpected bug. Trusted code runs the plan — the model never executes."""
    tables, primary, exts = base["tables"], base["primary"], base["exts"]

    partial_warning = None
    try:
        result, result_name, notes, render_ops = execute_multi(tables, primary, operations)
    except MultiStepError as exc:
        # A later step failed: keep the file reflecting the completed steps (PRD MS-b).
        result, result_name = exc.partial_result, exc.partial_name
        notes, render_ops = exc.notes, exc.format_ops
        done = exc.failed_step - 1
        partial_warning = (
            f"Step {exc.failed_step} couldn't be done: {exc.reason} "
            f"Your file reflects the {done} step{'s' if done != 1 else ''} that "
            "completed before it — fix that step and try again."
        )
    except OperationError as exc:
        return _error(str(exc), status=422)
    except Exception:
        return _error(_INTERNAL_ERROR, status=500)

    # Push the new state so the next instruction chains on it (and Retry/Edit can branch).
    if session_id:
        if isinstance(result, dict):
            new_state = {
                "tables": {**tables, **result},
                "primary": next(iter(result)),
                "exts": {**exts, **{k: "xlsx" for k in result}},
            }
        else:
            new_state = {
                "tables": {**tables, result_name: result},
                "primary": result_name,
                "exts": {**exts, result_name: exts.get(result_name, "xlsx")},
            }
        _push_state(session_id, new_state)

    biggest = max((len(t) for t in tables.values()), default=0)
    if biggest > 50_000:
        notes = [
            f"Heads up: this is a large file (~{biggest:,} rows) — it still processed, "
            "but big files can take a little longer."
        ] + notes

    try:
        if isinstance(result, dict):
            out_bytes, out_name, media_type = _serialize_workbook(result, result_name)
            row_count = sum(int(len(d)) for d in result.values())
        else:
            out_ext, upgrade_note = _output_ext(exts.get(result_name, "xlsx"), render_ops)
            if upgrade_note:
                notes = notes + [upgrade_note]
            out_bytes, out_name, media_type = _serialize(result, result_name, out_ext, render_ops)
            row_count = int(len(result))
    except Exception:
        return _error(_INTERNAL_ERROR, status=500)

    download_id = _store_result(out_bytes, out_name, media_type)
    inline_b64 = (
        base64.b64encode(out_bytes).decode("ascii")
        if len(out_bytes) <= _INLINE_MAX_BYTES else None
    )
    return JSONResponse(
        {
            "status": "ok",
            "session_id": session_id,
            "explanation": " ".join(notes) if notes else "No changes were needed.",
            "notes": notes,
            "formulas": _describe_formulas(render_ops),
            "row_count": row_count,
            "rows_before": int(len(tables[primary])) if primary in tables else None,
            "preview": _result_preview(result, result_name),
            "actions": list(dict.fromkeys(op.get("action") for op in operations)),
            "ai_title": ai_title,
            "partial": partial_warning is not None,
            "warning": partial_warning,
            "filename": out_name,
            "media_type": media_type,
            "file_size": len(out_bytes),
            "elapsed_ms": int((time.time() - started_at) * 1000),
            "download_id": download_id,
            "file_base64": inline_b64,
        }
    )


@app.post("/execute")
async def execute(
    session_id: str = Form(...),
    plan: str = Form(...),
    rewind: int = Form(-1),
) -> JSONResponse:
    """Phase 2 of the two-phase flow: the Hands ONLY. Run an already-approved plan
    (from /parse) on the session's data with NO model call. Same result shape as
    /process. The plan's columns/types are validated by the executor before it runs."""
    started_at = time.time()

    entry = _SESSIONS.get(session_id) if session_id else None
    if not entry or not entry.get("states"):
        return _error("Please upload a spreadsheet to start.", status=400)
    states = entry["states"]
    if 0 <= rewind < len(states):
        del states[rewind + 1:]  # Retry/Edit: branch from an earlier step
    base = states[-1]

    try:
        parsed = json.loads(plan)
    except Exception:
        return _error("That plan couldn't be read — please try running again.", status=400)
    operations = parsed.get("operations") or []
    if not operations:
        return _error("There's nothing to run.", status=400)
    ai_title = (parsed.get("title") or "").strip() or None

    return _run_operations(session_id, base, operations, ai_title, started_at)


@app.post("/process")
async def process(
    instruction: str = Form(...),
    session_id: str = Form(""),
    rewind: int = Form(-1),
    history: str = Form(""),
    files: list[UploadFile] = File(default=[]),
) -> JSONResponse:
    started_at = time.time()
    instruction = (instruction or "").strip()
    if not instruction:
        return _error("Please describe what you'd like done to the data.", status=400)

    # 1. Resolve the base working state. A session keeps a STACK of states:
    #    states[0] = the uploaded data, plus one more per successful step. `rewind`
    #    lets Retry/Edit re-run an earlier step by branching from the state just
    #    before it (dropping the now-stale later steps).
    if files:
        too_big = _too_big(files)
        if too_big:
            return _error(too_big, status=413)
        uploads = [(f.filename or "upload", await f.read()) for f in files]
        try:
            data = load_files(uploads)
        except ValueError as exc:
            return _error(str(exc), status=400)
        if session_id and session_id in _SESSIONS:
            cur = _SESSIONS[session_id]["states"][-1]
            base = {
                "tables": {**cur["tables"], **data.tables},
                "primary": data.primary,
                "exts": {**cur["exts"], **data.exts},
            }
        else:
            base = {"tables": dict(data.tables), "primary": data.primary, "exts": dict(data.exts)}
        if session_id:
            _remember_session(session_id, base)  # fresh upload resets step history (+ evicts old)
    elif session_id and session_id in _SESSIONS:
        states = _SESSIONS[session_id]["states"]
        if 0 <= rewind < len(states):
            del states[rewind + 1:]  # branch: drop steps at/after the rewind point
        base = states[-1]
    else:
        return _error(
            "Please upload a spreadsheet to start (or start a new session).", status=400
        )

    tables, primary, exts = base["tables"], base["primary"], base["exts"]

    # 2. Summarize all tables for the model so it can plan across files.
    structure = summarize_tables(tables, primary)

    # 3. Translate the instruction into an operation plan.
    try:
        plan = llm.parse_instruction(instruction, structure, history)
    except Exception as exc:
        # The Brain is unavailable (rate limit / quota / outage / parse error). Try a
        # deterministic fallback for simple commands so basic work still happens. The
        # fallback result is shown like any normal result (no "AI unavailable" notice —
        # that reads as a bad/uncertain experience to the user).
        unavailable = isinstance(exc, llm.ModelUnavailableError)
        key_missing = isinstance(exc, RuntimeError) and not unavailable
        if not (unavailable or key_missing):
            traceback.print_exc()  # log the real cause; never blame the instruction
        plan = fallback.parse(instruction, structure)
        if plan is None:
            if unavailable:
                return _error(str(exc), status=503)
            if key_missing:
                return _error(str(exc), status=500)
            return _error(
                "I couldn't reach the AI service to understand your request right now — "
                "this isn't a problem with your instruction. Please try again in a moment.",
                status=502,
            )

    clarification = plan.get("clarification")
    reply = plan.get("reply")
    ai_title = (plan.get("title") or "").strip() or None
    operations = plan.get("operations") or []
    if not operations:
        # No action to take: answer a data question, ask for clarity, or nudge.
        if reply:
            return JSONResponse({"status": "message", "message": reply})
        if clarification:
            return JSONResponse({"status": "clarify", "clarification": clarification})
        return JSONResponse(
            {
                "status": "message",
                "message": (
                    "I didn't understand that — try describing the task, e.g. "
                    '"sort by Revenue descending" or "remove duplicate rows".'
                ),
            }
        )

    # 4. Execute the plan across all tables.
    partial_warning = None
    try:
        result, result_name, notes, render_ops = execute_multi(tables, primary, operations)
    except MultiStepError as exc:
        # A later step failed: keep the file reflecting the steps that completed and
        # tell the user exactly which step failed and why (PRD MS-b).
        result, result_name = exc.partial_result, exc.partial_name
        notes, render_ops = exc.notes, exc.format_ops
        done = exc.failed_step - 1
        partial_warning = (
            f"Step {exc.failed_step} couldn't be done: {exc.reason} "
            f"Your file reflects the {done} step{'s' if done != 1 else ''} that "
            "completed before it — fix that step and try again."
        )
    except OperationError as exc:
        # Expected, user-facing problem (bad column, wrong type, …) — explain it.
        return _error(str(exc), status=422)
    except Exception:
        # Unexpected bug: never leak the exception/traceback to the user (1.14-d).
        return _error(_INTERNAL_ERROR, status=500)

    # 5. Push the new state so the next instruction chains on it (and so Retry/Edit
    #    can branch from any earlier step). Originals stay reachable for lookups/merges.
    if session_id:
        if isinstance(result, dict):
            new_state = {
                "tables": {**tables, **result},
                "primary": next(iter(result)),
                "exts": {**exts, **{k: "xlsx" for k in result}},
            }
        else:
            new_state = {
                "tables": {**tables, result_name: result},
                "primary": result_name,
                "exts": {**exts, result_name: exts.get(result_name, "xlsx")},
            }
        _push_state(session_id, new_state)

    # Large-file notice (PRD: big files still work, but tell the user).
    biggest = max((len(t) for t in tables.values()), default=0)
    if biggest > 50_000:
        notes = [
            f"Heads up: this is a large file (~{biggest:,} rows) — it still processed, "
            "but big files can take a little longer."
        ] + notes

    # 6. Serialize. The result is either one table or a multi-sheet workbook.
    try:
        if isinstance(result, dict):
            out_bytes, out_name, media_type = _serialize_workbook(result, result_name)
            row_count = sum(int(len(d)) for d in result.values())
        else:
            out_ext, upgrade_note = _output_ext(exts.get(result_name, "xlsx"), render_ops)
            if upgrade_note:
                notes = notes + [upgrade_note]
            out_bytes, out_name, media_type = _serialize(result, result_name, out_ext, render_ops)
            row_count = int(len(result))
    except Exception:  # saving the workbook failed unexpectedly — stay friendly (1.14-d)
        return _error(_INTERNAL_ERROR, status=500)

    # Always offer the file via a streamed download URL (small JSON, no browser OOM).
    # Only ALSO inline it as base64 when it's small enough to be cheap.
    download_id = _store_result(out_bytes, out_name, media_type)
    inline_b64 = (
        base64.b64encode(out_bytes).decode("ascii")
        if len(out_bytes) <= _INLINE_MAX_BYTES else None
    )
    return JSONResponse(
        {
            "status": "ok",
            "session_id": session_id,
            "explanation": " ".join(notes) if notes else "No changes were needed.",
            "notes": notes,
            "formulas": _describe_formulas(render_ops),
            "row_count": row_count,
            "rows_before": int(len(tables[primary])) if primary in tables else None,
            "preview": _result_preview(result, result_name),
            "actions": list(dict.fromkeys(op.get("action") for op in operations)),
            "ai_title": ai_title,
            "partial": partial_warning is not None,
            "warning": partial_warning,
            "filename": out_name,
            "media_type": media_type,
            "file_size": len(out_bytes),
            "elapsed_ms": int((time.time() - started_at) * 1000),
            "download_id": download_id,
            "file_base64": inline_b64,
        }
    )


def _result_preview(result, result_name: str, sample_rows: int = 8) -> list[dict]:
    """A compact, JSON-safe preview of the RESULT so the UI can SHOW the transformed
    data (not just offer a download). Same shape as /inspect's tables. For a
    multi-sheet workbook, previews each sheet."""
    frames = result if isinstance(result, dict) else {result_name: result}
    preview = []
    for name, df in frames.items():
        s = summarize_structure(df, sample_rows=sample_rows)
        preview.append({
            "name": name,
            "row_count": s["row_count"],
            "columns": s["columns"],
            "sample_rows": s["sample_rows"],
            "truncated": s["row_count"] > sample_rows,
        })
    return preview


def _describe_formulas(render_ops: list[dict]) -> list[str]:
    """Plain descriptions of any formulas/lookups added, for the UI's summary panel."""
    out = []
    for d in render_ops:
        if d.get("type") == "formula":
            out.append(f"{d['column']} = {d['formula']}")
        elif d.get("type") == "lookup":
            out.append(
                f"{d['new_column']} = look up '{d['return_column']}' from "
                f"'{d['source_name']}' matched on '{d['key_column']}'"
            )
    return out


def _output_ext(in_ext: str, render_ops: list[dict]) -> tuple[str, str | None]:
    """Choose the download extension. A .csv can't carry live formulas or formatting,
    so when an operation produced any render directive (formula/lookup/format/
    highlight) we upgrade the download to .xlsx — that's what makes the 1.13 promise
    ("formulas live, formatting intact") actually hold. Returns (ext, optional note)."""
    if in_ext == "csv" and render_ops:
        return "xlsx", (
            "Saved as .xlsx so the live formulas/formatting are preserved "
            "(a .csv file can't hold them)."
        )
    return in_ext, None


def _serialize_workbook(sheets: dict, base_name: str) -> tuple[bytes, str, str]:
    """Write several tables into ONE .xlsx, each on its own sheet/tab."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        taken: set[str] = set()
        for name, d in sheets.items():
            sheet_name = _safe_sheet_name(name, taken)
            taken.add(sheet_name)
            d.to_excel(writer, index=False, sheet_name=sheet_name)
            _disarm_injection(writer.sheets[sheet_name])
    stem = (base_name or "combined").rsplit(".", 1)[0]
    return (
        buf.getvalue(),
        f"{stem}_sumio.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _disarm_injection(ws) -> None:
    """Neutralize spreadsheet formula-injection from UPLOADED data.

    openpyxl writes any string that starts with '=' as a live formula, so a cell
    like '=HYPERLINK("http://evil")' in the user's file would execute when they open
    Sumio's output in Excel/Sheets. We force every such DATA cell to plain text (the
    visible value is unchanged). Call this right after writing data and BEFORE writing
    our own intentional formulas, so those are left live.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                cell.data_type = "s"


def _serialize(df, original_name: str, ext: str, render_ops=None) -> tuple[bytes, str, str]:
    render_ops = render_ops or []
    stem = (original_name or "result").rsplit(".", 1)[0]
    buf = io.BytesIO()
    if ext == "csv":
        # CSV has no styling/formulas; render directives are simply ignored.
        df.to_csv(buf, index=False)
        return buf.getvalue(), f"{stem}_sumio.csv", "text/csv"
    # xlsx via openpyxl
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        _disarm_injection(writer.sheets["Sheet1"])  # before our own formulas go in
        _apply_render(writer, "Sheet1", df, render_ops)
    return (
        buf.getvalue(),
        f"{stem}_sumio.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _apply_render(writer, main_name: str, df, render_ops: list[dict]) -> None:
    """Apply each render directive (formatting / live formula / highlight / lookup)
    to the workbook. Only meaningful for .xlsx output."""
    ws = writer.sheets[main_name]
    for directive in render_ops:
        kind = directive.get("type")
        if kind == "format":
            _apply_format(ws, df, directive)
        elif kind == "formula":
            _apply_formula(ws, df, directive)
        elif kind == "highlight":
            _apply_highlight(ws, df, directive)
        elif kind == "lookup":
            _apply_lookup(writer, ws, df, directive)


def _safe_sheet_name(base: str, taken: set[str]) -> str:
    """A valid, unique Excel sheet name (<=31 chars, no : \\ / ? * [ ])."""
    name = re.sub(r"[:\\/?*\[\]]", " ", str(base)).strip()[:28] or "Lookup"
    candidate = name
    i = 2
    while candidate in taken:
        candidate = f"{name} {i}"[:31]
        i += 1
    return candidate


def _apply_lookup(writer, main_ws, df, directive: dict) -> None:
    """Write the lookup source as its own sheet and a LIVE lookup formula into the
    column, so the result stays editable in Excel/Google Sheets.

    To keep the live file CONSISTENT with the computed preview, we add a hidden
    normalized-key helper column to the source sheet (trimmed, lowercased, numbers
    coerced to text) and match the lookup key against THAT — so case/space and
    number-vs-text differences match exactly as the preview did, not Excel's
    stricter exact match.
    """
    columns = list(df.columns)
    new_col = directive.get("new_column")
    key_col = directive.get("key_column")
    source_df = directive.get("source_df")
    skey = directive.get("source_key_column")
    sret = directive.get("return_column")
    if new_col not in columns or key_col not in columns or source_df is None:
        return
    src_cols = list(source_df.columns)
    if skey not in src_cols or sret not in src_cols or len(source_df) == 0:
        return

    sheet_name = _safe_sheet_name(directive.get("source_name") or "Lookup", set(writer.book.sheetnames))
    source_df.to_excel(writer, index=False, sheet_name=sheet_name)
    src_ws = writer.sheets[sheet_name]
    _disarm_injection(src_ws)  # the lookup source is uploaded data too

    n = len(source_df)
    q = sheet_name.replace("'", "''")
    sret_l = get_column_letter(src_cols.index(sret) + 1)
    key_l = get_column_letter(columns.index(key_col) + 1)
    target = columns.index(new_col) + 1

    # Write the hidden normalized-key helper column just past the source's columns.
    norm_keys = directive.get("source_norm_keys") or []
    helper_idx = len(src_cols) + 1
    helper_l = get_column_letter(helper_idx)
    src_ws.cell(row=1, column=helper_idx, value="_match_key")
    for i, k in enumerate(norm_keys[:n]):
        src_ws.cell(row=i + 2, column=helper_idx, value=k)
    src_ws.column_dimensions[helper_l].hidden = True

    match_range = f"'{q}'!${helper_l}$2:${helper_l}${n + 1}"
    return_range = f"'{q}'!${sret_l}$2:${sret_l}${n + 1}"
    for i in range(len(df)):
        r = i + 2
        formula = _lookup_formula(f"{key_l}{r}", match_range, return_range)
        main_ws.cell(row=r, column=target, value=formula)


def _norm_key_formula(key_cell: str) -> str:
    """Excel expression that normalizes a key cell the same way the backend does:
    coerce to text (so 123 == "123"), trim spaces, lowercase."""
    return f'LOWER(TRIM({key_cell}&""))'


def _lookup_formula(key_cell: str, match_range: str, return_range: str) -> str:
    """Build the lookup formula in the configured style, matching against the
    normalized helper column so live results equal the preview.

    INDEX/MATCH (default) works in every Excel version and Google Sheets; XLOOKUP
    is cleaner but needs Excel 2021/365 or Google Sheets.
    """
    key = _norm_key_formula(key_cell)
    if config.LOOKUP_STYLE == "xlookup":
        return f'=XLOOKUP({key},{match_range},{return_range},"Not found")'
    return f'=IFERROR(INDEX({return_range},MATCH({key},{match_range},0)),"Not found")'


def _apply_format(ws, df, directive: dict) -> None:
    columns = list(df.columns)
    if directive.get("bold_header"):
        for cell in ws[1]:
            cell.font = Font(bold=True)
    fmt = directive.get("format")
    if not fmt:
        return
    code = _number_format_code(
        fmt, directive.get("decimals"), directive.get("currency_symbol"),
        directive.get("date_format"),
    )
    for col in directive.get("columns") or []:
        if col not in columns:
            continue
        col_idx = columns.index(col) + 1  # openpyxl columns are 1-based
        for row in range(2, ws.max_row + 1):  # skip the header row
            ws.cell(row=row, column=col_idx).number_format = code


def _apply_formula(ws, df, directive: dict) -> None:
    """Write LIVE Excel formulas (e.g. =B2*C2) down the formula column, translating
    {ColumnName} placeholders into real cell references from the final layout."""
    columns = list(df.columns)
    name = directive.get("column")
    template = directive.get("formula") or ""
    if name not in columns:
        return
    referenced = _PLACEHOLDER.findall(template)
    # If any referenced column is gone (renamed/dropped later), keep the computed
    # values rather than writing a broken formula.
    if any(r not in columns for r in referenced):
        return
    target_idx = columns.index(name) + 1
    for i in range(len(df)):
        excel_row = i + 2  # row 1 is the header
        cell_formula = _PLACEHOLDER.sub(
            lambda m: f"{get_column_letter(columns.index(m.group(1)) + 1)}{excel_row}",
            template,
        )
        ws.cell(row=excel_row, column=target_idx, value="=" + cell_formula)


def _apply_highlight(ws, df, directive: dict) -> None:
    """Shade blank cells yellow without changing their (empty) value."""
    fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    columns = list(df.columns)
    for col in directive.get("columns") or []:
        if col not in columns:
            continue
        col_idx = columns.index(col) + 1
        series = df[col]
        for i in range(len(df)):
            v = series.iloc[i]
            if pd.isna(v) or (isinstance(v, str) and v.strip() == ""):
                ws.cell(row=i + 2, column=col_idx).fill = fill


# Friendly date-format names -> Excel number-format codes. Default is DD-MM-YYYY
# (Sameer's client-report style). The Brain may also pass a raw Excel code.
_DATE_FORMATS = {
    "dd-mm-yyyy": "dd-mm-yyyy",
    "mm-dd-yyyy": "mm-dd-yyyy",
    "yyyy-mm-dd": "yyyy-mm-dd",
    "dd/mm/yyyy": "dd/mm/yyyy",
    "mm/dd/yyyy": "mm/dd/yyyy",
    "dd-mmm-yyyy": "dd-mmm-yyyy",        # 09-Jun-2026
    "d mmmm yyyy": "d mmmm yyyy",        # 9 June 2026
    "mmmm d, yyyy": 'mmmm d", "yyyy',    # June 9, 2026
}


def _date_format_code(date_format) -> str:
    if not date_format:
        return "dd-mm-yyyy"
    key = str(date_format).strip().lower()
    if key in _DATE_FORMATS:
        return _DATE_FORMATS[key]
    # Looks like a raw Excel date code (only d/m/y, separators, spaces) -> pass through.
    if re.fullmatch(r"[dmyDMY/\-\. ,]+", str(date_format).strip()):
        return str(date_format).strip()
    return "dd-mm-yyyy"


def _number_format_code(fmt: str, decimals, symbol, date_format=None) -> str:
    """Translate a friendly format name into an Excel number-format code."""
    d = decimals if isinstance(decimals, int) and decimals >= 0 else 2
    dec = "." + "0" * d if d > 0 else ""
    if fmt == "currency":
        return f'"{symbol or "₹"}"#,##0{dec}'
    if fmt == "percent":
        return f"0{dec}%"
    if fmt == "number":
        return f"#,##0{dec}"
    if fmt == "date":
        return _date_format_code(date_format)
    return "General"


# Shown for any UNEXPECTED failure, so a bug never reaches the user as a stack
# trace or raw error code (PRD 1.14-d / "no raw error codes ever reach the user").
_INTERNAL_ERROR = (
    "Something went wrong on our side while processing your file — "
    "please try again, or rephrase your instruction."
)


def _too_big(files) -> str | None:
    """Friendly message if the combined upload exceeds the size limit, else None.
    Guards against a single huge upload exhausting server memory."""
    limit = config.MAX_UPLOAD_MB * 1024 * 1024
    total = sum((getattr(f, "size", None) or 0) for f in files)
    if total > limit:
        return (
            f"That upload is too large (~{total / 1024 / 1024:.0f} MB). "
            f"Please keep files under {config.MAX_UPLOAD_MB} MB."
        )
    return None


def _remember_session(session_id: str, state: dict) -> None:
    """Store a fresh session state, bounding both the number of sessions and the
    per-session undo stack so memory can't grow without limit."""
    _SESSIONS[session_id] = {"states": [state]}
    while len(_SESSIONS) > config.MAX_SESSIONS:
        _SESSIONS.pop(next(iter(_SESSIONS)))  # evict the oldest (dicts keep order)


def _push_state(session_id: str, state: dict) -> None:
    """Append a new step to a session, trimming the OLDEST steps past the cap while
    always keeping states[0] (the original upload, needed for rewind + lookups)."""
    entry = _SESSIONS.setdefault(session_id, {"states": []})
    states = entry["states"]
    states.append(state)
    if len(states) > config.MAX_STATES:
        entry["states"] = [states[0]] + states[-(config.MAX_STATES - 1):]


def _error(message: str, status: int) -> JSONResponse:
    return JSONResponse({"status": "error", "error": message}, status_code=status)
