"""Feature 1.9 — Lookup (VLOOKUP / XLOOKUP). Verifies 1.9-a..f + acceptance criteria
and edges, including that the saved .xlsx contains a LIVE lookup formula that matches
the source on its OWN sheet, and that the live result stays consistent with the
computed preview (case/space/number-vs-text normalization).

Run from backend:  .venv\\Scripts\\python.exe test_1_9.py
"""
from __future__ import annotations

import io

import openpyxl
import pandas as pd

from app import config
from app.executor import OperationError, execute_plan, _norm_key
from app.main import _serialize, _describe_formulas

passed = 0
failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
    else:
        failed += 1
        print(f"  FAIL  {name}  {detail}")


def run(op, df, sheets):
    """Run one lookup op; lookup reaches other tables via the sheets namespace."""
    return execute_plan(df, [op], sheets=sheets)


def lookup_op(**kw):
    base = {"action": "lookup"}
    base.update(kw)
    return base


def load_xlsx(out, render):
    return openpyxl.load_workbook(io.BytesIO(_serialize(out, "orders", "xlsx", render)[0]))


print("FEATURE 1.9 — lookup (VLOOKUP / XLOOKUP)\n")

# --------------------------------------------------------------------------- #
# 1.9-a  Simple cross-sheet lookup -> correct values fetched for each key
# --------------------------------------------------------------------------- #
orders = pd.DataFrame({"CustID": [1, 2, 3], "Amount": [100, 200, 300]})
people = pd.DataFrame({"ID": [1, 2, 3], "Name": ["Asha", "Rohan", "Sameer"]})
sheets = {"Orders": orders, "People": people}
out, notes, render = run(
    lookup_op(key_column="CustID", source_sheet="People",
              source_key_column="ID", return_column="Name"),
    orders, sheets,
)
check("1.9-a fetched values", list(out["Name"]) == ["Asha", "Rohan", "Sameer"], str(list(out.get("Name"))))
check("1.9-a new column added", "Name" in out.columns)
check("1.9-a note mentions matched", "3 of 3" in notes[0], notes[0])
check("1.9-a emits lookup directive", any(d.get("type") == "lookup" for d in render))

# acceptance: cross-sheet — source written as its OWN sheet, with a LIVE formula
wb = load_xlsx(out, render)
check("1.9-a source on its own sheet", "People" in wb.sheetnames, str(wb.sheetnames))
main_ws = wb["Sheet1"]
name_col = list(out.columns).index("Name") + 1
cell = main_ws.cell(row=2, column=name_col).value
check("1.9-a live formula written", isinstance(cell, str) and cell.startswith("="), repr(cell))
check("1.9-a formula references source sheet", isinstance(cell, str) and "People" in cell, repr(cell))
check("1.9-a formula has Not found fallback", isinstance(cell, str) and '"Not found"' in cell, repr(cell))

# --------------------------------------------------------------------------- #
# 1.9-b  Key with no match -> "Not found" marker, no crash (value + formula)
# --------------------------------------------------------------------------- #
orders2 = pd.DataFrame({"CustID": [1, 99], "Amount": [10, 20]})
out, notes, render = run(
    lookup_op(key_column="CustID", source_sheet="People",
              source_key_column="ID", return_column="Name"),
    orders2, {"Orders": orders2, "People": people},
)
check("1.9-b matched value", out["Name"].iloc[0] == "Asha", str(out["Name"].iloc[0]))
check("1.9-b unmatched -> 'Not found'", out["Name"].iloc[1] == "Not found", str(out["Name"].iloc[1]))
check("1.9-b note reports 1 of 2", "1 of 2" in notes[0], notes[0])

# --------------------------------------------------------------------------- #
# 1.9-c  Duplicate keys in source -> first match used; user informed
# --------------------------------------------------------------------------- #
dup_src = pd.DataFrame({"ID": [1, 1, 2], "Name": ["First", "Second", "Other"]})
ord3 = pd.DataFrame({"CustID": [1, 2]})
out, notes, render = run(
    lookup_op(key_column="CustID", source_sheet="Src",
              source_key_column="ID", return_column="Name"),
    ord3, {"Orders": ord3, "Src": dup_src},
)
check("1.9-c first match used", out["Name"].iloc[0] == "First", str(out["Name"].iloc[0]))
check("1.9-c user informed of dupes", "duplicate" in notes[0].lower() and "first" in notes[0].lower(), notes[0])

# --------------------------------------------------------------------------- #
# 1.9-d  Key types differ (number vs text "123") -> matched + clearly flagged
# --------------------------------------------------------------------------- #
ord4 = pd.DataFrame({"CustID": ["123", "456"]})         # keys are TEXT
src4 = pd.DataFrame({"ID": [123, 456], "Name": ["NumA", "NumB"]})  # keys are NUMBERS
out, notes, render = run(
    lookup_op(key_column="CustID", source_sheet="Src",
              source_key_column="ID", return_column="Name"),
    ord4, {"Orders": ord4, "Src": src4},
)
check("1.9-d number/text matched", list(out["Name"]) == ["NumA", "NumB"], str(list(out["Name"])))
check("1.9-d behavior flagged", "number-vs-text" in notes[0] or "only after ignoring" in notes[0], notes[0])

# --------------------------------------------------------------------------- #
# 1.9-e  Source sheet doesn't exist -> friendly error
# --------------------------------------------------------------------------- #
try:
    run(lookup_op(key_column="CustID", source_sheet="Missing",
                  source_key_column="ID", return_column="Name"),
        orders, {"Orders": orders})
    check("1.9-e missing sheet caught", False, "no error raised")
except OperationError as e:
    check("1.9-e friendly missing-sheet msg", "Missing" in str(e) and "Available" in str(e), str(e))

# --------------------------------------------------------------------------- #
# 1.9-f  Case/space differences in keys -> matched (dedupe-style default), stated
# --------------------------------------------------------------------------- #
ord5 = pd.DataFrame({"Email": ["A@X.com", " bob@x.com "]})
src5 = pd.DataFrame({"mail": ["a@x.com", "bob@x.com"], "Name": ["Ann", "Bob"]})
out, notes, render = run(
    lookup_op(key_column="Email", source_sheet="Src",
              source_key_column="mail", return_column="Name"),
    ord5, {"Orders": ord5, "Src": src5},
)
check("1.9-f case/space matched", list(out["Name"]) == ["Ann", "Bob"], str(list(out["Name"])))
check("1.9-f behavior stated", "ignoring case" in notes[0] or "surrounding spaces" in notes[0], notes[0])

# --------------------------------------------------------------------------- #
# Live file stays CONSISTENT with the preview: hidden normalized helper column
# on the source sheet + formula matches against it (so 1.9-d/f match live too).
# --------------------------------------------------------------------------- #
wb = load_xlsx(out, render)
src_ws = wb["Src"]
helper_idx = len(src5.columns) + 1
helper_header = src_ws.cell(row=1, column=helper_idx).value
check("helper column header is _match_key", helper_header == "_match_key", repr(helper_header))
helper_vals = [src_ws.cell(row=i + 2, column=helper_idx).value for i in range(len(src5))]
check("helper column holds normalized keys", helper_vals == list(_norm_key(src5["mail"])), str(helper_vals))
from openpyxl.utils import get_column_letter
check("helper column hidden", src_ws.column_dimensions[get_column_letter(helper_idx)].hidden is True)
# the live formula normalizes the key cell the same way (LOWER + TRIM + &"")
mws = wb["Sheet1"]
f = mws.cell(row=2, column=list(out.columns).index("Name") + 1).value
check("formula normalizes key cell", isinstance(f, str) and "LOWER(TRIM(" in f and '&\"\"' in f, repr(f))
check("formula matches helper column", isinstance(f, str) and "_match_key" not in f and "$" in f, repr(f))

# --------------------------------------------------------------------------- #
# Acceptance / edge cases
# --------------------------------------------------------------------------- #

# new_column defaults to the return column name when omitted
out, _, _ = run(
    lookup_op(key_column="CustID", source_sheet="People",
              source_key_column="ID", return_column="Name"),
    orders, sheets,
)
check("edge: new_column defaults to return column", "Name" in out.columns)

# explicit new_column name is honored
out, _, _ = run(
    lookup_op(key_column="CustID", source_sheet="People", source_key_column="ID",
              return_column="Name", new_column="Customer Name"),
    orders, sheets,
)
check("edge: explicit new_column honored", "Customer Name" in out.columns, str(list(out.columns)))

# missing required fields -> friendly error
try:
    run(lookup_op(key_column="CustID", source_sheet="People"), orders, sheets)
    check("edge: missing fields caught", False, "no error")
except OperationError as e:
    check("edge: missing fields friendly", "needs" in str(e).lower(), str(e))

# key column not in the current sheet -> friendly error
try:
    run(lookup_op(key_column="Nope", source_sheet="People",
                  source_key_column="ID", return_column="Name"), orders, sheets)
    check("edge: bad key column caught", False, "no error")
except OperationError as e:
    check("edge: bad key column friendly", "Nope" in str(e), str(e))

# return/source-key column not in the SOURCE sheet -> friendly error listing columns
try:
    run(lookup_op(key_column="CustID", source_sheet="People",
                  source_key_column="ID", return_column="Salary"), orders, sheets)
    check("edge: bad return column caught", False, "no error")
except OperationError as e:
    check("edge: bad return column friendly", "Salary" in str(e) and "has no column" in str(e), str(e))

# blank key in the current sheet -> "Not found" (blank source keys never match)
ord_blank = pd.DataFrame({"CustID": [1, None]})
src_blank = pd.DataFrame({"ID": [1, None], "Name": ["Asha", "Ghost"]})
out, _, _ = run(
    lookup_op(key_column="CustID", source_sheet="Src",
              source_key_column="ID", return_column="Name"),
    ord_blank, {"Orders": ord_blank, "Src": src_blank},
)
check("edge: blank key -> Not found", out["Name"].iloc[1] == "Not found", str(out["Name"].iloc[1]))

# CSV output: lookup values are written (no formulas in CSV) and don't crash
ord_csv = pd.DataFrame({"CustID": [1, 2]})
out, _, render = run(
    lookup_op(key_column="CustID", source_sheet="People",
              source_key_column="ID", return_column="Name"),
    ord_csv, {"Orders": ord_csv, "People": people},
)
csv_bytes = _serialize(out, "orders", "csv", render)[0]
check("edge: CSV has computed values", b"Asha" in csv_bytes and b"Rohan" in csv_bytes)

# UI summary panel describes the lookup in plain language
desc = _describe_formulas(render)
check("UI: lookup described for summary panel",
      any("look up" in d and "Name" in d for d in desc), str(desc))

# XLOOKUP style honored when configured
_old_style = config.LOOKUP_STYLE
try:
    config.LOOKUP_STYLE = "xlookup"
    out, _, render = run(
        lookup_op(key_column="CustID", source_sheet="People",
                  source_key_column="ID", return_column="Name"),
        orders, sheets,
    )
    wb = load_xlsx(out, render)
    f = wb["Sheet1"].cell(row=2, column=list(out.columns).index("Name") + 1).value
    check("edge: XLOOKUP style", isinstance(f, str) and f.startswith("=XLOOKUP("), repr(f))
finally:
    config.LOOKUP_STYLE = _old_style

# index_match (default) is used otherwise
out, _, render = run(
    lookup_op(key_column="CustID", source_sheet="People",
              source_key_column="ID", return_column="Name"),
    orders, sheets,
)
wb = load_xlsx(out, render)
f = wb["Sheet1"].cell(row=2, column=list(out.columns).index("Name") + 1).value
check("edge: index_match default", isinstance(f, str) and f.startswith("=IFERROR(INDEX("), repr(f))

# empty source sheet -> no live formula written, but values computed (Not found) safely
ord_e = pd.DataFrame({"CustID": [1]})
empty_src = pd.DataFrame({"ID": [], "Name": []})
out, _, render = run(
    lookup_op(key_column="CustID", source_sheet="Src",
              source_key_column="ID", return_column="Name"),
    ord_e, {"Orders": ord_e, "Src": empty_src},
)
check("edge: empty source -> Not found, no crash", out["Name"].iloc[0] == "Not found", str(out["Name"].iloc[0]))
wb = load_xlsx(out, render)  # should not raise
check("edge: empty source serializes", "Sheet1" in wb.sheetnames)

print(f"\n{passed} passed, {failed} failed.")
raise SystemExit(1 if failed else 0)
