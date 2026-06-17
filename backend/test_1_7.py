"""Feature 1.7 — Handle missing values (fill / drop / flag). Verifies 1.7-a..e + edges
at the executor/serializer level. 1.7-f (decline statistical fill) is Brain-driven.

Run from backend:  .venv\\Scripts\\python.exe test_1_7.py
"""
from __future__ import annotations

import io

import openpyxl
import pandas as pd

from app.executor import OperationError, execute_plan
from app.main import _serialize

passed = 0
failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
    else:
        failed += 1
        print(f"  FAIL  {name}  {detail}")


def run(op, df):
    return execute_plan(df, [op])


print("FEATURE 1.7 — missing values\n")

# 1.7-a fill blanks with a value; count reported
out, notes, _ = run({"action": "fill_missing", "columns": ["R"], "fill_value": "Unknown"}, pd.DataFrame({"R": ["x", None, ""]}))
check("1.7-a fill replaces blanks", (out["R"] == "Unknown").sum() == 2, str(list(out["R"])))
check("1.7-a count reported", "Filled 2 blank cells" in notes[0], notes[0])
# fill numeric keeps it numeric
out, _, _ = run({"action": "fill_missing", "columns": ["n"], "fill_value": "0"}, pd.DataFrame({"n": [1.0, None]}))
check("1.7-a numeric fill stays number", out["n"].iloc[1] == 0 and out["n"].iloc[1] != "0")

# 1.7-b drop rows with blanks; count reported
out, notes, _ = run({"action": "drop_missing", "columns": ["R"]}, pd.DataFrame({"R": ["x", None, ""], "k": [1, 2, 3]}))
check("1.7-b drop rows with blanks", len(out) == 1 and out["R"].iloc[0] == "x")
check("1.7-b count reported", "Removed 2 row" in notes[0], notes[0])

# 1.7-c flag blanks: data unchanged + highlight directive; serialized xlsx shades blanks
df = pd.DataFrame({"R": ["x", None, ""]})
out, notes, render = run({"action": "flag_missing", "columns": ["R"]}, df)
check("1.7-c flag leaves data unchanged", list(out["R"].fillna("")) == ["x", "", ""])
check("1.7-c emits highlight directive", any(d.get("type") == "highlight" for d in render))
ws = openpyxl.load_workbook(io.BytesIO(_serialize(out, "o", "xlsx", render)[0])).active
check("1.7-c blank cell highlighted", ws.cell(row=3, column=1).fill.fill_type == "solid")
check("1.7-c non-blank not highlighted", ws.cell(row=2, column=1).fill.fill_type in (None, "none"))

# 1.7-d no blanks present -> "No missing values found"
check("1.7-d fill none", "No missing values found" in run({"action": "fill_missing", "columns": ["R"], "fill_value": "U"}, pd.DataFrame({"R": ["a", "b"]}))[1][0])
check("1.7-d drop none", "No missing values found" in run({"action": "drop_missing", "columns": ["R"]}, pd.DataFrame({"R": ["a", "b"]}))[1][0])
check("1.7-d flag none", "No missing values found" in run({"action": "flag_missing", "columns": ["R"]}, pd.DataFrame({"R": ["a", "b"]}))[1][0])

# simple fills: previous value (ffill) and next value (bfill)
out, notes, _ = run({"action": "fill_missing", "columns": ["R"], "fill_method": "previous"}, pd.DataFrame({"R": ["a", None, "", "b", None]}))
check("fill previous (ffill)", list(out["R"]) == ["a", "a", "a", "b", "b"], str(list(out["R"])))
check("fill previous count", "using the previous value" in notes[0], notes[0])
out, _, _ = run({"action": "fill_missing", "columns": ["R"], "fill_method": "next"}, pd.DataFrame({"R": ["a", None, "b", None]}))
check("fill next (bfill)", list(out["R"].iloc[:3]) == ["a", "b", "b"] and pd.isna(out["R"].iloc[3]), str(list(out["R"])))
# leading blank with ffill stays blank (no previous), no crash
out, _, _ = run({"action": "fill_missing", "columns": ["R"], "fill_method": "previous"}, pd.DataFrame({"R": [None, "x", None]}))
check("fill previous leaves leading blank", pd.isna(out["R"].iloc[0]) and list(out["R"].iloc[1:]) == ["x", "x"], str(list(out["R"])))

# 1.7-e entire column blank -> handled without crash
out, notes, _ = run({"action": "fill_missing", "columns": ["R"], "fill_value": "U"}, pd.DataFrame({"R": [None, None, ""]}))
check("1.7-e fill whole-blank column", (out["R"] == "U").all() and "Filled 3" in notes[0])
out, notes, _ = run({"action": "drop_missing", "columns": ["R"]}, pd.DataFrame({"R": [None, None]}))
check("1.7-e drop whole-blank column (empty result, no crash)", len(out) == 0)

# edges: fill without value -> error; non-existent column -> error
try:
    run({"action": "fill_missing", "columns": ["R"]}, pd.DataFrame({"R": [None]}))
    check("edge: fill needs a value", False, "no error")
except OperationError:
    check("edge: fill needs a value", True)
try:
    run({"action": "drop_missing", "columns": ["Ghost"]}, pd.DataFrame({"R": [1]}))
    check("edge: non-existent column caught", False)
except OperationError as e:
    check("edge: non-existent column caught", "Ghost" in str(e))

print(f"\n{passed} passed, {failed} failed.")
raise SystemExit(1 if failed else 0)
